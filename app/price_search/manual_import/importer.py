from __future__ import annotations

import csv
from dataclasses import asdict
from decimal import Decimal
from difflib import SequenceMatcher
from pathlib import Path
from typing import Any

from openpyxl import load_workbook
from sqlalchemy import select
from sqlalchemy.orm import Session

from app.models import MarketOffer, Purchase, PurchaseItem
from app.price_search.base import MarketOfferCandidate
from app.price_search.manual_import.schemas import ManualOfferRow, OfferImportResult
from app.price_search.normalization import (
    normalize_delivery_price,
    normalize_price,
    normalize_quantity,
    normalize_region,
    normalize_title,
    normalize_url,
)
from app.price_search.relevance import calculate_offer_relevance


EXPECTED_COLUMNS = [
    "purchase_external_id",
    "position_external_id",
    "item_name",
    "offer_title",
    "offer_url",
    "seller_name",
    "region",
    "unit_price",
    "available_quantity",
    "delivery_price",
    "delivery_days",
    "is_relevant",
    "relevance_score",
    "comment",
]


def import_offers_from_file(session: Session, file_path: Path) -> OfferImportResult:
    result = OfferImportResult(file_path=str(file_path))

    rows = _load_rows(file_path)
    result.total_rows = len(rows)

    purchases_by_external = {
        purchase.external_id: purchase for purchase in session.scalars(select(Purchase)).all()
    }
    items = session.scalars(select(PurchaseItem).where(PurchaseItem.purchase_id.is_not(None))).all()

    for idx, raw in enumerate(rows, start=2):
        try:
            parsed = _parse_manual_row(raw)
            if parsed is None:
                result.skipped_count += 1
                continue

            purchase, item = _resolve_purchase_and_item(
                parsed=parsed,
                purchases_by_external=purchases_by_external,
                all_items=items,
            )

            candidate = _candidate_from_manual_row(parsed, item_id=item.id if item else None)
            if item is not None:
                relevance = calculate_offer_relevance(item, candidate)
                candidate.is_relevant = relevance.is_relevant if parsed.is_relevant is None else parsed.is_relevant
                candidate.relevance_score = parsed.relevance_score if parsed.relevance_score is not None else relevance.score
                candidate.risk_flags = sorted(set(candidate.risk_flags + relevance.risk_flags))
            else:
                if parsed.is_relevant is not None:
                    candidate.is_relevant = parsed.is_relevant
                if parsed.relevance_score is not None:
                    candidate.relevance_score = parsed.relevance_score

            offer = MarketOffer(
                provider="manual_import",
                source="manual",
                purchase_id=purchase.id if purchase else None,
                purchase_item_id=item.id if item else None,
                purchase_external_id=parsed.purchase_external_id,
                position_external_id=parsed.position_external_id,
                item_name=parsed.item_name,
                offer_title=parsed.offer_title,
                offer_url=candidate.url,
                seller_name=parsed.seller_name,
                supplier_name=parsed.seller_name or "unknown",
                region=candidate.region,
                region_code=_region_code_from_text(candidate.region),
                unit_price=candidate.unit_price,
                available_quantity=int(candidate.available_quantity),
                delivery_price=candidate.delivery_price,
                delivery_days=candidate.delivery_days,
                effective_unit_price=_effective_unit_price(
                    unit_price=candidate.unit_price,
                    available_quantity=candidate.available_quantity,
                    delivery_price=candidate.delivery_price,
                ),
                relevance_score=Decimal(str(candidate.relevance_score)),
                is_relevant=bool(candidate.is_relevant),
                risk_flags=candidate.risk_flags,
                comment=parsed.comment,
                raw_payload=_json_safe(asdict(parsed)),
            )
            session.add(offer)
            result.imported_count += 1
        except Exception as exc:  # noqa: BLE001
            result.error_count += 1
            result.errors.append(f"line {idx}: {exc}")

    session.commit()
    return result


def _load_rows(file_path: Path) -> list[dict[str, Any]]:
    suffix = file_path.suffix.lower()
    if suffix == ".csv":
        with file_path.open("r", encoding="utf-8-sig", newline="") as fp:
            reader = csv.DictReader(fp)
            return [dict(row) for row in reader]

    if suffix in {".xlsx", ".xlsm"}:
        wb = load_workbook(file_path)
        ws = wb.active
        header = [str(cell.value).strip() if cell.value is not None else "" for cell in ws[1]]
        rows: list[dict[str, Any]] = []
        for row in ws.iter_rows(min_row=2, values_only=True):
            payload = {header[idx]: value for idx, value in enumerate(row) if idx < len(header)}
            rows.append(payload)
        return rows

    raise ValueError("Unsupported file format. Use .csv or .xlsx")


def _parse_manual_row(raw: dict[str, Any]) -> ManualOfferRow | None:
    if not any(raw.get(key) not in (None, "") for key in EXPECTED_COLUMNS):
        return None

    item_name = str(raw.get("item_name") or "").strip()
    offer_title = str(raw.get("offer_title") or "").strip()
    if not item_name or not offer_title:
        raise ValueError("item_name and offer_title are required")

    unit_price = normalize_price(raw.get("unit_price"))
    if unit_price is None or unit_price <= 0:
        raise ValueError("unit_price must be positive")

    quantity, quantity_flags = normalize_quantity(raw.get("available_quantity"))
    delivery_price, delivery_flags = normalize_delivery_price(raw.get("delivery_price"))
    region, region_flags = normalize_region(raw.get("region"))

    risk_flags = quantity_flags + delivery_flags + region_flags

    is_relevant_raw = raw.get("is_relevant")
    is_relevant = None
    if is_relevant_raw is not None and str(is_relevant_raw).strip() != "":
        is_relevant = str(is_relevant_raw).strip().lower() in {"1", "true", "yes", "y"}

    relevance_score_raw = raw.get("relevance_score")
    relevance_score = float(relevance_score_raw) if relevance_score_raw not in (None, "") else None

    delivery_days_raw = raw.get("delivery_days")
    delivery_days = int(delivery_days_raw) if delivery_days_raw not in (None, "") else None

    return ManualOfferRow(
        purchase_external_id=_clean_str(raw.get("purchase_external_id")),
        position_external_id=_clean_str(raw.get("position_external_id")),
        item_name=item_name,
        offer_title=offer_title,
        offer_url=normalize_url(_clean_str(raw.get("offer_url"))),
        seller_name=_clean_str(raw.get("seller_name")),
        region=region,
        unit_price=unit_price,
        available_quantity=quantity,
        delivery_price=delivery_price,
        delivery_days=delivery_days,
        is_relevant=is_relevant,
        relevance_score=relevance_score,
        comment=_clean_str(raw.get("comment")),
        raw_payload={**raw, "risk_flags": sorted(set(risk_flags))},
    )


def _resolve_purchase_and_item(
    parsed: ManualOfferRow,
    purchases_by_external: dict[str, Purchase],
    all_items: list[PurchaseItem],
) -> tuple[Purchase | None, PurchaseItem | None]:
    purchase = purchases_by_external.get(parsed.purchase_external_id) if parsed.purchase_external_id else None

    if purchase and parsed.position_external_id:
        for item in all_items:
            if item.purchase_id == purchase.id and item.position_external_id == parsed.position_external_id:
                return purchase, item

    if purchase and parsed.item_name:
        purchase_items = [item for item in all_items if item.purchase_id == purchase.id]
        item = _find_best_item_match(parsed.item_name, purchase_items)
        return purchase, item

    if parsed.item_name:
        item = _find_best_item_match(parsed.item_name, all_items)
        if item:
            return purchase, item

    return purchase, None


def _find_best_item_match(item_name: str, candidates: list[PurchaseItem]) -> PurchaseItem | None:
    normalized_target = normalize_title(item_name)
    best_item: PurchaseItem | None = None
    best_ratio = 0.0

    for candidate in candidates:
        ratio = SequenceMatcher(None, normalized_target, normalize_title(candidate.item_name)).ratio()
        if ratio > best_ratio:
            best_ratio = ratio
            best_item = candidate

    if best_ratio >= 0.72:
        return best_item
    return None


def _candidate_from_manual_row(parsed: ManualOfferRow, item_id: int | None) -> MarketOfferCandidate:
    return MarketOfferCandidate(
        provider="manual_import",
        purchase_item_id=item_id,
        title=parsed.offer_title,
        url=parsed.offer_url,
        seller_name=parsed.seller_name,
        region=parsed.region,
        unit_price=parsed.unit_price,
        available_quantity=parsed.available_quantity,
        delivery_price=parsed.delivery_price,
        delivery_days=parsed.delivery_days,
        is_relevant=True,
        relevance_score=parsed.relevance_score if parsed.relevance_score is not None else 1.0,
        raw_payload=parsed.raw_payload,
        risk_flags=list((parsed.raw_payload or {}).get("risk_flags", [])),
        item_name=parsed.item_name,
        comment=parsed.comment,
    )


def _effective_unit_price(unit_price: Decimal, available_quantity: Decimal, delivery_price: Decimal | None) -> Decimal:
    qty = max(available_quantity, Decimal("1"))
    delivery = delivery_price or Decimal("0")
    return unit_price + (delivery / qty)


def _clean_str(value: Any) -> str | None:
    if value is None:
        return None
    text = str(value).strip()
    return text or None


def _region_code_from_text(region: str | None) -> str | None:
    if not region:
        return None
    lowered = region.lower()
    if "моск" in lowered and "обл" in lowered:
        return "50"
    if "моск" in lowered:
        return "77"
    return None


def _json_safe(value):
    if isinstance(value, Decimal):
        return str(value)
    if isinstance(value, dict):
        return {key: _json_safe(item) for key, item in value.items()}
    if isinstance(value, list):
        return [_json_safe(item) for item in value]
    if isinstance(value, tuple):
        return [_json_safe(item) for item in value]
    return value
