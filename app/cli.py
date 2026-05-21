from __future__ import annotations

import json
import os
import socket
import subprocess
import sys
from pathlib import Path
from typing import Any

import requests
import typer
import uvicorn
from openpyxl import load_workbook
from sqlalchemy import text

from app.browser import BrowserSessionManager
from app.connectors.eat.api_client import EatApiClient
from app.connectors.eat.browser_fallback import EatBrowserFallback
from app.connectors.eat.parser import parse_purchase_payload as parse_eat_purchase_payload
from app.connectors.mos_portal.api_client import MosPortalApiClient
from app.connectors.mos_portal.browser_fallback import MosPortalBrowserFallback
from app.connectors.mos_portal.parser import parse_purchase_payload as parse_mos_purchase_payload
from app.config import ConfigValidationError, get_required_no_proxy_hosts, get_settings
from app.db import SessionLocal
from app.demo.seed_demo_data import reset_demo_data, seed_demo_data
from app.price_search.manual_import.importer import import_offers_from_file
from app.reports.daily_digest import generate_daily_digest
from app.scheduler import run_scheduler, scheduler_status
from app.scoring.strategy import activate_strategy, create_default_strategies, get_active_strategy, list_strategies
from app.security.redaction import redact_mapping
from app.services.backup_service import backup_file_stats, cleanup_backups, create_backup, list_backups, restore_backup
from app.services.decision_service import DecisionService
from app.services.data_validation_service import DataValidationService
from app.services.excel_export_service import export_to_excel
from app.services.financial_check_service import FinancialCheckService
from app.services.fixture_loader import load_fixtures
from app.services.import_service import ImportService
from app.services.item_attribute_service import ItemAttributeService
from app.services.job_service import create_job_run, mark_failed, mark_running, mark_success
from app.services.offer_template_service import export_offer_template
from app.services.rematch_service import rematch_offers
from app.services.supplier_service import SupplierService
from app.services.task_runner import run_calculate_task, run_evaluate_task, run_export_excel_task, run_import_offers_task, run_parse_task, run_search_prices_task
from app.services.user_service import UserService
from app.services.watchlist_service import WatchlistService
from app.utils.time import utc_now
from app.utils.migrations import (
    SQLITE_DEMO_MIGRATION_MODE,
    database_backend,
    ensure_runtime_directories,
    get_current_revision,
    get_head_revision,
    init_database_for_settings,
    migration_mode_for_settings,
    migrations_are_up_to_date,
    upgrade_head,
)
from app.version import get_version

cli = typer.Typer(help="Tender small volume calculator CLI")
supplier_cli = typer.Typer(help="Supplier management")
strategy_cli = typer.Typer(help="Participation strategies")
watchlist_cli = typer.Typer(help="Watchlist management")
user_cli = typer.Typer(help="Users and roles management")
db_cli = typer.Typer(help="Database migration workflow")
demo_cli = typer.Typer(help="Demo mode and seed data")
backup_cli = typer.Typer(help="Backup and restore")
parse_cli = typer.Typer(help="Parse/import purchases")
price_search_cli = typer.Typer(help="Price search operations")
diagnostics_cli = typer.Typer(help="Diagnostics and health")
scheduler_cli = typer.Typer(help="Scheduler operations")
env_cli = typer.Typer(help="Environment file helpers")
browser_cli = typer.Typer(help="Browser/Playwright helpers")
cli.add_typer(supplier_cli, name="supplier")
cli.add_typer(strategy_cli, name="strategy")
cli.add_typer(watchlist_cli, name="watchlist")
cli.add_typer(user_cli, name="user")
cli.add_typer(user_cli, name="users")
cli.add_typer(db_cli, name="db")
cli.add_typer(demo_cli, name="demo")
cli.add_typer(backup_cli, name="backup")
cli.add_typer(parse_cli, name="parse-group")
cli.add_typer(price_search_cli, name="price-search")
cli.add_typer(diagnostics_cli, name="diagnostics")
cli.add_typer(scheduler_cli, name="scheduler")
cli.add_typer(env_cli, name="env")
cli.add_typer(browser_cli, name="browser")


@cli.command("init-db")
def init_db() -> None:
    settings = get_settings()
    ensure_runtime_directories(settings)
    mode = init_database_for_settings(settings)
    if mode == SQLITE_DEMO_MIGRATION_MODE:
        typer.echo("Initialized SQLite demo database using SQLAlchemy metadata.create_all.")
        return
    typer.echo("Database migrations applied via Alembic.")


@env_cli.command("init")
def env_init_command(
    mode: str = typer.Option("development", "--mode", help="development | demo | production"),
    output: Path = typer.Option(Path(".env"), "--output"),
) -> None:
    source = Path(".env.example")
    if not source.exists():
        raise typer.BadParameter(".env.example not found")

    target_mode = mode.strip().lower()
    if target_mode not in {"development", "demo", "production"}:
        raise typer.BadParameter("mode must be development, demo or production")

    if output.exists():
        typer.echo(f"{output} already exists, not overwriting.")
        return

    content = source.read_text(encoding="utf-8")
    lines = content.splitlines()
    values = _env_to_dict(lines)

    if target_mode == "demo":
        values.update(
            {
                "APP_MODE": "demo",
                "DEMO_DATA_ENABLED": "true",
                "REAL_NETWORK_ENABLED": "false",
                "REAL_RUN_MODE": "false",
                "APP_ENV": "demo",
                "DASHBOARD_SECRET_KEY": "qa-demo-secret",
                "DASHBOARD_ALLOWED_HOSTS": "127.0.0.1,localhost,testserver",
                "DATABASE_URL": "sqlite+pysqlite:///data/qa_demo.db",
                "USE_PROXY": "false",
            }
        )
    elif target_mode == "production":
        values.update(
            {
                "APP_MODE": "production",
                "DEMO_DATA_ENABLED": "false",
                "REAL_NETWORK_ENABLED": "true",
                "REAL_RUN_MODE": "true",
                "DASHBOARD_AUTH_ENABLED": "true",
                "DASHBOARD_SECRET_KEY": "CHANGE_ME_TO_STRONG_RANDOM_SECRET",
                "DATABASE_URL": "postgresql+psycopg://postgres:postgres@localhost:5432/tender_small_volume",
            }
        )

    rendered = _render_env_lines(lines, values)
    header = [f"# Generated by: python -m app.cli env init --mode {target_mode}"]
    if target_mode == "production":
        header.extend(
            [
                "# IMPORTANT: replace DASHBOARD_SECRET_KEY and DATABASE_URL before first production run.",
                "# Keep DASHBOARD_AUTH_ENABLED=true in production.",
            ]
        )
    output.write_text("\n".join(header + [""] + rendered) + "\n", encoding="utf-8")
    typer.echo(f"Environment file created: {output}")


@cli.command("load-fixtures")
def load_fixtures_command(
    purchases_path: Path = Path("fixtures/sample_purchases.json"),
    offers_path: Path = Path("fixtures/sample_market_offers.json"),
    reset: bool = True,
) -> None:
    with SessionLocal() as session:
        load_fixtures(session, purchases_path=purchases_path, offers_path=offers_path, reset=reset)
    typer.echo("Fixtures loaded")


@cli.command("parse")
def parse_command(
    source: str = typer.Option("mos_portal", "--source", help="mos_portal | eat | all"),
    limit: int | None = typer.Option(None, "--limit", min=1),
    dry_run: bool = typer.Option(False, "--dry-run"),
    save_raw: bool = typer.Option(False, "--save-raw"),
    status: str = typer.Option("Прием предложений", "--status"),
) -> None:
    with SessionLocal() as session:
        job, report = run_parse_task(
            session=session,
            source=source,
            status=status,
            limit=limit,
            dry_run=dry_run,
            save_raw=save_raw,
            owner="cli",
        )
        job_id, job_status = job.id, job.status

    typer.echo(f"JobRun: {job_id} ({job_status})")
    if report is None:
        typer.echo("Skipped due to lock")
        return

    for result in report.results:
        typer.echo(f"Source: {result.source}")
        typer.echo(f"  found: {result.found_count}")
        typer.echo(f"  filtered: {result.filtered_count}")
        typer.echo(f"  created: {result.created_count}")
        typer.echo(f"  updated: {result.updated_count}")
        typer.echo(f"  skipped: {result.skipped_count}")
        typer.echo(f"  errors: {result.error_count}")
        for message in (result.errors or [])[:20]:
            typer.echo(f"    - {message}")


@cli.command("validate-data")
def validate_data_command(
    purchase_id: int | None = typer.Option(None, "--purchase-id"),
) -> None:
    with SessionLocal() as session:
        summary = DataValidationService(session).validate(purchase_id=purchase_id)
    typer.echo(f"checked_purchases: {summary.checked_purchases}")
    typer.echo(f"quality_high: {summary.high_quality}")
    typer.echo(f"quality_medium: {summary.medium_quality}")
    typer.echo(f"quality_low: {summary.low_quality}")
    typer.echo(f"warnings_count: {summary.warnings_count}")


@cli.command("real-source-check")
def real_source_check_command(
    source: str = typer.Option(..., "--source", help="mos_portal | eat"),
    limit: int = typer.Option(5, "--limit", min=1, max=20),
    save: bool = typer.Option(False, "--save"),
    status: str = typer.Option("Прием предложений", "--status"),
) -> None:
    settings = get_settings()
    source_norm = source.strip().lower()
    if source_norm not in {"mos_portal", "eat"}:
        raise typer.BadParameter("source must be mos_portal or eat")

    if settings.app_mode == "demo" or not settings.real_network_enabled:
        typer.echo("Real network is disabled in demo mode.")
        return

    missing_no_proxy = [host for host in _required_hosts_for_source(source_norm) if host not in settings.no_proxy]
    if missing_no_proxy:
        typer.echo(f"NO_PROXY missing required hosts for {source_norm}: {', '.join(missing_no_proxy)}")

    check = _run_real_source_probe(source_norm, status=status, limit=limit)
    typer.echo(f"source: {source_norm}")
    typer.echo(f"connector_status: {check['connector_status']}")
    typer.echo(f"api_ok: {check['api_ok']}")
    typer.echo(f"browser_fallback_used: {check['browser_fallback_used']}")
    typer.echo(f"auth_required: {check['auth_required']}")
    typer.echo(f"captcha_or_blocked: {check['captcha_or_blocked']}")
    typer.echo(f"fetched_raw_records: {check['raw_records_count']}")
    typer.echo(f"parsed_purchases: {len(check['parsed'])}")
    metrics = _parsed_validation_metrics(check["parsed"])
    typer.echo(f"FOUND: {metrics['found']} purchases")
    typer.echo(f"VALID: {metrics['valid']}")
    typer.echo("MISSING_FIELDS:")
    if metrics["missing_fields"]:
        for key, count in sorted(metrics["missing_fields"].items()):
            typer.echo(f"- {key} ({count})")
    else:
        typer.echo("- none")
    for warning in check["warnings"][:10]:
        typer.echo(f"warning: {warning}")
    for error in check["errors"][:10]:
        typer.echo(f"error: {error}")

    _print_parsed_sample(check["parsed"], limit=min(limit, 5))

    if save and check["parsed"]:
        with SessionLocal() as session:
            result = ImportService(session=session, save_raw=False).import_purchases(
                source=source_norm,
                parsed_purchases=check["parsed"],
                dry_run=False,
                required_status=status,
            )
        typer.echo(
            "saved: found={found} created={created} updated={updated} filtered={filtered} errors={errors}".format(
                found=result.found_count,
                created=result.created_count,
                updated=result.updated_count,
                filtered=result.filtered_count,
                errors=result.error_count,
            )
        )
    elif save:
        typer.echo("Nothing to save: no parsed purchases")
    else:
        typer.echo("dry-run only: add --save to persist parsed purchases")


@cli.command("export-offer-template")
def export_offer_template_command(file: Path = typer.Option(..., "--file")) -> None:
    output = export_offer_template(file)
    typer.echo(f"Template exported: {output}")


@cli.command("import-offers")
def import_offers_command(file: Path = typer.Option(..., "--file")) -> None:
    with SessionLocal() as session:
        job, result = run_import_offers_task(session=session, file_path=file, owner="cli", importer_fn=import_offers_from_file)
        job_id, job_status = job.id, job.status

    typer.echo(f"JobRun: {job_id} ({job_status})")
    if result is None:
        typer.echo("Skipped due to lock")
        return

    typer.echo(f"File: {result.file_path}")
    typer.echo(f"  total_rows: {result.total_rows}")
    typer.echo(f"  imported: {result.imported_count}")
    typer.echo(f"  skipped: {result.skipped_count}")
    typer.echo(f"  errors: {result.error_count}")
    for message in (result.errors or [])[:20]:
        typer.echo(f"    - {message}")


@cli.command("search-prices")
def search_prices_command(
    mode: str = typer.Option("manual", "--mode", help="stub | manual | yandex"),
    limit: int | None = typer.Option(None, "--limit", min=1),
    purchase_id: int | None = typer.Option(None, "--purchase-id"),
    item_id: int | None = typer.Option(None, "--item-id"),
) -> None:
    settings = get_settings()
    if mode.strip().lower() == "yandex" and (settings.app_mode == "demo" or not settings.real_network_enabled):
        typer.echo("Real network is disabled in demo mode.")
        raise typer.Exit(code=1)

    with SessionLocal() as session:
        job, result = run_search_prices_task(
            session=session,
            mode=mode,
            limit=limit,
            purchase_id=purchase_id,
            item_id=item_id,
            owner="cli",
        )
        job_id, job_status = job.id, job.status

    typer.echo(f"JobRun: {job_id} ({job_status})")
    if result is None:
        typer.echo("Skipped due to lock")
        return

    typer.echo(f"Mode: {result.mode}")
    typer.echo(f"  processed_items: {result.processed_items}")
    typer.echo(f"  created_offers: {result.created_offers}")
    typer.echo(f"  needs_manual_items: {result.needs_manual_items}")
    typer.echo(f"  errors: {len(result.errors)}")
    for message in result.errors[:20]:
        typer.echo(f"    - {message}")


@cli.command("calculate")
def calculate_command() -> None:
    with SessionLocal() as session:
        job, result = run_calculate_task(session=session, purchase_id=None, owner="cli")
        job_id, job_status = job.id, job.status
    typer.echo(f"JobRun: {job_id} ({job_status})")
    typer.echo(f"Calculated purchases: {result}")


@cli.command("financial-check")
def financial_check_command(
    purchase_id: int | None = typer.Option(None, "--purchase-id"),
) -> None:
    with SessionLocal() as session:
        summary = FinancialCheckService(session).check(purchase_id=purchase_id)
    typer.echo(f"checked_purchases: {summary.checked_purchases}")
    typer.echo(f"ok: {summary.ok_count}")
    typer.echo(f"warnings: {summary.warning_count}")
    typer.echo(f"errors: {summary.error_count}")
    for row in summary.rows[:20]:
        if row.status == "ok":
            continue
        typer.echo(
            f"purchase_id={row.purchase_id} status={row.status} "
            f"errors={'; '.join(row.errors) if row.errors else '-'} "
            f"warnings={'; '.join(row.warnings) if row.warnings else '-'}"
        )


@cli.command("recalculate")
def recalculate_command(
    purchase_id: int | None = typer.Option(None, "--purchase-id"),
    all: bool = typer.Option(False, "--all"),
    with_attributes: bool = typer.Option(False, "--with-attributes"),
) -> None:
    if not all and purchase_id is None:
        raise typer.BadParameter("Use --purchase-id or --all")

    with SessionLocal() as session:
        if with_attributes:
            attr_service = ItemAttributeService(session)
            if all:
                extracted = attr_service.extract_for_all_missing()
                typer.echo(f"Extracted missing attributes: {extracted}")
            else:
                attr_service.extract_for_purchase(purchase_id)
                typer.echo(f"Attributes refreshed for purchase: {purchase_id}")

        if all:
            job, result = run_calculate_task(session=session, purchase_id=None, owner="cli")
            typer.echo(f"JobRun: {job.id} ({job.status})")
            typer.echo(f"Recalculated all purchases: {result}")
        else:
            job, _ = run_calculate_task(session=session, purchase_id=purchase_id, owner="cli")
            typer.echo(f"JobRun: {job.id} ({job.status})")
            typer.echo(f"Recalculated purchase: {purchase_id}")


@cli.command("extract-attributes")
def extract_attributes_command(
    all: bool = typer.Option(False, "--all"),
    purchase_id: int | None = typer.Option(None, "--purchase-id"),
    item_id: int | None = typer.Option(None, "--item-id"),
) -> None:
    if not all and purchase_id is None and item_id is None:
        raise typer.BadParameter("Use --all, --purchase-id or --item-id")

    with SessionLocal() as session:
        service = ItemAttributeService(session)
        if all:
            count = service.extract_for_all_missing()
            typer.echo(f"Extracted attributes for missing items: {count}")
            return
        if purchase_id is not None:
            count = service.extract_for_purchase(purchase_id)
            typer.echo(f"Extracted attributes for purchase {purchase_id}: {count}")
            return
        if item_id is not None:
            row = service.refresh_attributes(item_id)
            typer.echo(f"Extracted attributes for item {item_id}: category={row.category}, brand={row.brand}, article={row.article}")


@cli.command("rematch-offers")
def rematch_offers_command(
    all: bool = typer.Option(False, "--all"),
    purchase_id: int | None = typer.Option(None, "--purchase-id"),
    item_id: int | None = typer.Option(None, "--item-id"),
) -> None:
    if not all and purchase_id is None and item_id is None:
        raise typer.BadParameter("Use --all, --purchase-id or --item-id")
    with SessionLocal() as session:
        updated = rematch_offers(session=session, purchase_id=None if all else purchase_id, item_id=item_id)
    typer.echo(f"Offers rematched: {updated}")


@cli.command("evaluate")
def evaluate_command(
    purchase_id: int | None = typer.Option(None, "--purchase-id"),
) -> None:
    with SessionLocal() as session:
        job, result = run_evaluate_task(session=session, purchase_id=purchase_id, owner="cli")
        job_id, job_status = job.id, job.status
    typer.echo(f"JobRun: {job_id} ({job_status})")
    typer.echo(f"Evaluate result: {result}")


@cli.command("opportunities")
def opportunities_command(limit: int = typer.Option(20, "--limit", min=1, max=200)) -> None:
    with SessionLocal() as session:
        rows = DecisionService(session).get_top_opportunities(limit=limit)
    for row in rows:
        typer.echo(
            f"purchase_id={row.purchase_id} decision={row.decision} risk={row.risk_level} "
            f"score={row.score_total} next_action={row.next_action}"
        )


@cli.command("daily-digest")
def daily_digest_command(send: bool = typer.Option(False, "--send")) -> None:
    with SessionLocal() as session:
        digest = generate_daily_digest(session, send=send)
    typer.echo(f"generated_at: {digest.generated_at}")
    typer.echo(f"new_purchases_24h: {digest.new_purchases_24h}")
    typer.echo(f"calculated_24h: {digest.calculated_24h}")
    typer.echo(f"strong_recommend: {digest.strong_recommend_count}")
    typer.echo(f"recommend: {digest.recommend_count}")
    typer.echo(f"needs_review: {digest.needs_review_count}")


@cli.command("export-decisions")
def export_decisions_command(file: Path = typer.Option(Path("exports/decisions.xlsx"), "--file")) -> None:
    with SessionLocal() as session:
        output = export_to_excel(session, output_path=file)
    typer.echo(f"Decisions exported: {output}")


@cli.command("export-excel")
def export_excel_command(output_path: Path = Path("exports/tender_small_volume_export.xlsx")) -> None:
    with SessionLocal() as session:
        job, result = run_export_excel_task(session=session, output_path=output_path, owner="cli")
        job_id, job_status = job.id, job.status
    typer.echo(f"JobRun: {job_id} ({job_status})")
    typer.echo(f"Excel exported: {result}")


@cli.command("validate-export")
def validate_export_command(file: Path = typer.Option(..., "--file")) -> None:
    if not file.exists():
        typer.echo(f"Export file not found: {file}")
        raise typer.Exit(code=1)

    try:
        wb = load_workbook(file, data_only=True)
    except Exception as exc:  # noqa: BLE001
        typer.echo(f"Failed to open Excel file: {exc}")
        raise typer.Exit(code=1)

    required_sheets = {"Summary", "Purchases", "Items", "Offers", "Calculations"}
    missing = sorted(required_sheets - set(wb.sheetnames))
    if missing:
        typer.echo(f"Missing required sheets: {', '.join(missing)}")
        raise typer.Exit(code=1)

    checks = {
        "Summary": wb["Summary"].max_row > 1,
        "Purchases": wb["Purchases"].max_row > 1,
        "Items": wb["Items"].max_row > 1,
        "Offers": wb["Offers"].max_row > 1,
    }
    failed = [name for name, ok in checks.items() if not ok]
    for name, ok in checks.items():
        typer.echo(f"{name}: {'ok' if ok else 'empty'}")
    if failed:
        typer.echo(f"Validation failed. Empty sheets: {', '.join(failed)}")
        raise typer.Exit(code=1)

    typer.echo("Export validation passed")


@cli.command("run-dashboard")
def run_dashboard(host: str = "0.0.0.0", port: int = 8000) -> None:
    uvicorn.run("app.main:app", host=host, port=port, reload=False)


@cli.command("dashboard-snapshot")
def dashboard_snapshot_command(
    base_url: str = typer.Option("http://127.0.0.1:8000", "--base-url"),
    output_dir: Path = typer.Option(Path("artifacts/screenshots"), "--output-dir"),
    username: str | None = typer.Option(None, "--username"),
    password: str | None = typer.Option(None, "--password"),
) -> None:
    try:
        from playwright.sync_api import sync_playwright  # type: ignore
    except Exception:
        typer.echo("Playwright is not installed. Run python -m app.cli browser install")
        raise typer.Exit(code=1)

    pages = {
        "home": "/",
        "jobs": "/jobs",
        "diagnostics": "/diagnostics",
        "risks": "/risks",
        "watchlist": "/watchlist",
        "daily_report": "/reports/daily",
    }
    with SessionLocal() as session:
        first_purchase_id = session.execute(text("SELECT id FROM purchases ORDER BY id LIMIT 1")).scalar()
    if first_purchase_id is not None:
        pages["purchase_detail"] = f"/purchases/{first_purchase_id}"

    output_dir.mkdir(parents=True, exist_ok=True)
    settings = get_settings()
    try:
        with sync_playwright() as p:
            browser = p.chromium.launch(headless=True)
            context = browser.new_context()
            page = context.new_page()
            if settings.dashboard_auth_enabled:
                if not username or not password:
                    typer.echo("Dashboard auth is enabled. Provide --username and --password for snapshots.")
                    raise typer.Exit(code=1)
                page.goto(f"{base_url}/login", wait_until="domcontentloaded")
                page.fill("input[name='username']", username)
                page.fill("input[name='password']", password)
                page.click("button[type='submit']")
                page.wait_for_timeout(500)

            for name, route in pages.items():
                page.goto(f"{base_url}{route}", wait_until="domcontentloaded")
                path = output_dir / f"{name}.png"
                page.screenshot(path=str(path), full_page=True)
                typer.echo(f"saved: {path}")

            context.close()
            browser.close()
    except Exception as exc:  # noqa: BLE001
        typer.echo(f"dashboard-snapshot failed: {exc}")
        typer.echo("Ensure dashboard is running: python -m app.cli run-dashboard")
        raise typer.Exit(code=1)


@cli.command("run-scheduler")
def run_scheduler_command() -> None:
    run_scheduler()


@cli.command("scheduler-status")
def scheduler_status_command() -> None:
    status = scheduler_status()
    typer.echo(f"enabled: {status.enabled}")
    typer.echo(f"timezone: {status.timezone}")
    typer.echo(f"configured_jobs: {', '.join(status.configured_jobs)}")


@cli.command("browser-login")
def browser_login_command(source: str = typer.Option(..., "--source", help="mos_portal | eat")) -> None:
    try:
        manager = BrowserSessionManager()
        result = manager.login(source)
    except Exception as exc:  # noqa: BLE001
        typer.echo(f"browser-login failed: {exc}")
        raise typer.Exit(code=1)
    typer.echo(f"source: {result.source}")
    typer.echo(f"ok: {result.ok}")
    typer.echo(f"message: {result.message}")
    typer.echo(f"storage_state: {result.storage_state_path}")


@cli.command("browser-check")
def browser_check_command(source: str = typer.Option(..., "--source", help="mos_portal | eat")) -> None:
    try:
        manager = BrowserSessionManager()
        result = manager.check(source)
    except Exception as exc:  # noqa: BLE001
        typer.echo(f"browser-check failed: {exc}")
        raise typer.Exit(code=1)
    typer.echo(f"source: {result.source}")
    typer.echo(f"ok: {result.ok}")
    typer.echo(f"message: {result.message}")
    typer.echo(f"storage_state: {result.storage_state_path}")


@browser_cli.command("install")
def browser_install_command() -> None:
    try:
        import playwright  # noqa: F401
    except Exception:
        typer.echo("Playwright package is not installed. Run: pip install playwright")
        raise typer.Exit(code=1)

    cmd = [sys.executable, "-m", "playwright", "install", "chromium"]
    typer.echo(f"Running: {' '.join(cmd)}")
    try:
        subprocess.run(cmd, check=True)
    except subprocess.CalledProcessError as exc:
        typer.echo(f"Playwright browser install failed: {exc}")
        raise typer.Exit(code=1)
    typer.echo("Playwright chromium installed.")


@browser_cli.command("doctor")
def browser_doctor_command() -> None:
    settings = get_settings()
    issues: list[str] = []

    try:
        from playwright.sync_api import sync_playwright  # type: ignore
        typer.echo("playwright package: installed")
    except Exception:
        typer.echo("playwright package: missing")
        typer.echo("Run: pip install playwright")
        raise typer.Exit(code=1)

    chromium_path = None
    try:
        with sync_playwright() as p:
            chromium_path = p.chromium.executable_path
            browser = p.chromium.launch(headless=True)
            browser.close()
    except Exception as exc:  # noqa: BLE001
        issues.append(f"Chromium is not available: {exc}")

    if chromium_path:
        typer.echo(f"chromium executable: {chromium_path}")

    settings.browser_storage_state_dir.mkdir(parents=True, exist_ok=True)
    typer.echo(f"BROWSER_STORAGE_STATE_DIR: {settings.browser_storage_state_dir}")
    typer.echo(f"MOS_PORTAL_STORAGE_STATE: {settings.mos_portal_storage_state}")
    typer.echo(f"EAT_STORAGE_STATE: {settings.eat_storage_state}")

    if settings.app_mode == "demo" and not settings.real_network_enabled:
        typer.echo("real network is disabled in demo mode; browser-check/login should be used for manual setup only.")

    if issues:
        for issue in issues:
            typer.echo(f"- {issue}")
        typer.echo("Run: python -m app.cli browser install")
        raise typer.Exit(code=1)

    typer.echo("Browser doctor passed")


@cli.command("healthcheck")
def healthcheck_command() -> None:
    settings = get_settings()
    database = "ok"
    with SessionLocal() as session:
        try:
            session.execute(text("SELECT 1"))
        except Exception:
            database = "fail"

    scheduler_state = "not_running" if settings.scheduler_enabled else "unknown"
    typer.echo(
        json_dumps(
            {
                "status": "ok" if database == "ok" else "degraded",
                "database": database,
                "version": get_version(),
                "time": utc_now().isoformat(),
                "scheduler": scheduler_state,
            }
        )
    )


@cli.command("doctor")
def doctor_command() -> None:
    issues: list[str] = []
    warnings: list[str] = []

    typer.echo(f"Python version: {sys.version.split()[0]}")
    if sys.version_info < (3, 11):
        issues.append("Python 3.11+ required")

    env_exists = Path(".env").exists()
    typer.echo(f".env exists: {env_exists}")
    if not env_exists:
        warnings.append(".env file is missing")

    try:
        settings = get_settings()
        typer.echo("config validation: ok")
    except ConfigValidationError as exc:
        issues.append(f"config validation failed: {exc}")
        settings = None

    if settings is not None:
        if not settings.database_url.strip():
            issues.append("DATABASE_URL is empty")

        backend = database_backend(settings.database_url)
        migration_mode = migration_mode_for_settings(settings)
        typer.echo(f"app mode: {settings.app_mode}")
        typer.echo(f"real run mode: {settings.real_run_mode}")
        typer.echo(f"database backend: {backend}")
        typer.echo(f"migrations mode: {migration_mode}")

        for d in ensure_runtime_directories(settings):
            typer.echo(f"dir ok: {d}")

        missing_no_proxy = [host for host in get_required_no_proxy_hosts() if host not in settings.no_proxy]
        if missing_no_proxy:
            issues.append(f"NO_PROXY missing: {', '.join(sorted(missing_no_proxy))}")

        with SessionLocal() as session:
            try:
                session.execute(text("SELECT 1"))
                typer.echo("database connection: ok")
            except Exception as exc:  # noqa: BLE001
                issues.append(f"database connection failed: {exc}")

            if backend == "sqlite":
                if settings.app_mode in {"demo", "development"}:
                    warnings.append("SQLite is intended only for demo/development smoke runs.")
                else:
                    issues.append("SQLite is not supported outside demo/development mode. Use PostgreSQL.")
            elif backend == "postgresql":
                try:
                    current_revision = session.execute(text("SELECT version_num FROM alembic_version")).scalar()
                    typer.echo(f"alembic revision: {current_revision}")
                except Exception:
                    warnings.append("could not read alembic_version (migrations may be missing)")
                try:
                    head_revision = get_head_revision()
                    typer.echo(f"alembic head: {head_revision}")
                    if not migrations_are_up_to_date():
                        issues.append("database migrations are not up-to-date")
                except Exception as exc:  # noqa: BLE001
                    warnings.append(f"could not verify migrations: {exc}")
            else:
                warnings.append(f"Unknown database backend from DATABASE_URL: {settings.database_url}")

        try:
            import playwright  # noqa: F401
            typer.echo("playwright: installed")
        except Exception:
            warnings.append("playwright not installed")

        manager = BrowserSessionManager()
        typer.echo(f"browser state mos_portal exists: {manager.state_exists('mos_portal')}")
        typer.echo(f"browser state eat exists: {manager.state_exists('eat')}")

        port_busy = _is_port_open("127.0.0.1", 8000)
        typer.echo(f"dashboard port 8000 open: {port_busy}")

        if settings.notifications_enabled:
            if not settings.telegram_bot_token or not settings.telegram_chat_id:
                issues.append("notifications enabled but telegram settings are incomplete")
        if settings.app_mode == "production":
            if not settings.dashboard_auth_enabled:
                issues.append("production mode requires DASHBOARD_AUTH_ENABLED=true")
            if settings.dashboard_secret_key == "change-me":
                issues.append("production mode requires custom DASHBOARD_SECRET_KEY")
        if settings.app_mode == "demo" and settings.real_network_enabled:
            warnings.append("demo mode with REAL_NETWORK_ENABLED=true: set false for safe demo runs")
        if settings.real_run_mode and not settings.real_network_enabled:
            warnings.append("REAL_RUN_MODE=true but REAL_NETWORK_ENABLED=false: real import step will be skipped")

        redacted = redact_mapping(
            {
                "database_url": settings.database_url,
                "telegram_bot_token": settings.telegram_bot_token or "",
                "http_proxy": settings.http_proxy or "",
                "https_proxy": settings.https_proxy or "",
            }
        )
        typer.echo(f"redacted config snapshot: {redacted}")

    if warnings:
        typer.echo("Warnings:")
        for item in warnings:
            typer.echo(f"  - {item}")

    if issues:
        typer.echo("Issues:")
        for item in issues:
            typer.echo(f"  - {item}")
        raise typer.Exit(code=1)

    typer.echo("Doctor check passed")


@cli.command("backup-db")
def backup_db_command() -> None:
    output_file = create_backup()
    typer.echo(f"Backup created: {output_file}")


@cli.command("list-backups")
def list_backups_command() -> None:
    files = list_backups()
    if not files:
        typer.echo("No backups found")
        return
    for file in files:
        typer.echo(str(file))


@cli.command("backup-list")
def backup_list_alias_command() -> None:
    list_backups_command()


@cli.command("restore-db")
def restore_db_command(
    file: Path = typer.Option(..., "--file"),
    yes: bool = typer.Option(False, "--yes", help="Confirm restore operation"),
) -> None:
    restored, safety = restore_backup(file, yes=yes)
    typer.echo(f"Restore complete: {restored}")
    typer.echo(f"Safety backup created: {safety}")


@cli.command("backup-cleanup")
def backup_cleanup_command(keep_last: int = typer.Option(10, "--keep-last", min=1)) -> None:
    result = cleanup_backups(keep_last=keep_last)
    typer.echo(f"Backups cleanup: kept={result['kept']} removed={result['removed']} total_before={result['total_before']}")


@backup_cli.command("create")
def backup_create_group_command() -> None:
    backup_db_command()


@backup_cli.command("list")
def backup_list_group_command() -> None:
    list_backups_command()


@backup_cli.command("restore")
def backup_restore_group_command(
    file: Path = typer.Option(..., "--file"),
    yes: bool = typer.Option(False, "--yes"),
) -> None:
    restore_db_command(file=file, yes=yes)


@backup_cli.command("cleanup")
def backup_cleanup_group_command(keep_last: int = typer.Option(10, "--keep-last", min=1)) -> None:
    backup_cleanup_command(keep_last=keep_last)


@cli.command("run-real-pipeline")
def run_real_pipeline_command(
    dry_run: bool = typer.Option(False, "--dry-run"),
    source: str = typer.Option("mos_portal", "--source"),
    parse_limit: int | None = typer.Option(20, "--parse-limit", min=1),
    parse_status: str = typer.Option("Прием предложений", "--parse-status"),
    output_path: Path = typer.Option(Path("exports/real_pipeline_export.xlsx"), "--output-path"),
) -> None:
    settings = get_settings()
    ensure_runtime_directories(settings)
    init_database_for_settings(settings)

    with SessionLocal() as session:
        job = create_job_run(
            session=session,
            job_type="run_real_pipeline",
            source=source,
            params_json={
                "dry_run": dry_run,
                "source": source,
                "parse_limit": parse_limit,
                "parse_status": parse_status,
                "output_path": str(output_path),
            },
        )
        try:
            mark_running(session, job.id)
            parse_report = None
            if settings.real_network_enabled:
                _job, parse_report = run_parse_task(
                    session=session,
                    source=source,
                    status=parse_status,
                    limit=parse_limit,
                    dry_run=False,
                    save_raw=True,
                    owner="run_real_pipeline",
                )
            else:
                typer.echo("parse skipped: REAL_NETWORK_ENABLED=false")

            validation = DataValidationService(session).validate()
            extracted = ItemAttributeService(session).extract_for_all_missing()
            typer.echo("price-search step skipped in safe real pipeline (no yandex auto-search).")
            _calc_job, calc_count = run_calculate_task(session=session, purchase_id=None, owner="run_real_pipeline")
            financial = FinancialCheckService(session).check()
            _eval_job, eval_count = run_evaluate_task(session=session, purchase_id=None, owner="run_real_pipeline")
            _export_job, export_file = run_export_excel_task(session=session, output_path=output_path, owner="run_real_pipeline")

            missing_prices = session.execute(
                text(
                    "SELECT COUNT(DISTINCT purchase_id) FROM item_cost_calculations "
                    "WHERE status IN ('no_relevant_offers', 'needs_manual_price_search')"
                )
            ).scalar() or 0

            result_json = {
                "parse_total_found": getattr(parse_report, "total_found", 0) if parse_report else 0,
                "validated": validation.checked_purchases,
                "data_quality_low": validation.low_quality,
                "attributes_extracted": extracted,
                "calculated": calc_count,
                "financial_errors": financial.error_count,
                "evaluated": eval_count,
                "missing_prices_purchases": missing_prices,
                "export_file": str(export_file),
                "dry_run": dry_run,
            }
            mark_success(session, job.id, result_json=result_json)
        except Exception as exc:  # noqa: BLE001
            mark_failed(session, job.id, error_message=str(exc))
            raise

    typer.echo("run-real-pipeline completed")
    typer.echo(f"dry_run: {dry_run}")
    typer.echo(f"export: {output_path}")
    typer.echo("Use dashboard Next Actions to review missing prices / low data quality / approvals.")


@cli.command("run-all")
def run_all(
    output_path: Path = Path("exports/tender_small_volume_export.xlsx"),
    source: str = "mos_portal",
    parse_limit: int | None = None,
    parse_status: str = "Прием предложений",
    price_search_mode: str | None = None,
    with_dashboard: bool = False,
    host: str = "0.0.0.0",
    port: int = 8000,
) -> None:
    settings = get_settings()
    ensure_runtime_directories(settings)
    init_database_for_settings(settings)
    mode = price_search_mode or settings.run_all_price_search_mode

    with SessionLocal() as session:
        job = create_job_run(
            session=session,
            job_type="run_all",
            source=source,
            params_json={
                "output_path": str(output_path),
                "source": source,
                "parse_limit": parse_limit,
                "parse_status": parse_status,
                "price_search_mode": mode,
            },
        )
        try:
            mark_running(session, job.id)
            _, parse_report = run_parse_task(
                session=session,
                source=source,
                status=parse_status,
                limit=parse_limit,
                dry_run=False,
                save_raw=True,
                owner="run_all",
            )
            attr_service = ItemAttributeService(session)
            if source == "all":
                attr_service.extract_for_all_missing()
            else:
                # Refresh attributes for items parsed in current source.
                attr_service.extract_for_all_missing()
            _, search_result = run_search_prices_task(
                session=session,
                mode=mode,
                limit=parse_limit,
                purchase_id=None,
                item_id=None,
                owner="run_all",
            )
            _, calc_result = run_calculate_task(session=session, purchase_id=None, owner="run_all")
            _, eval_result = run_evaluate_task(session=session, purchase_id=None, owner="run_all")
            _, export_result = run_export_excel_task(session=session, output_path=output_path, owner="run_all")

            mark_success(
                session,
                job.id,
                result_json={
                    "parse": _safe_obj(parse_report),
                    "search": _safe_obj(search_result),
                    "calculate": _safe_obj(calc_result),
                    "evaluate": _safe_obj(eval_result),
                    "export": str(export_result),
                },
            )
        except Exception as exc:  # noqa: BLE001
            mark_failed(session, job.id, error_message=str(exc))
            raise

    typer.echo("run-all finished")
    if with_dashboard:
        uvicorn.run("app.main:app", host=host, port=port, reload=False)


@supplier_cli.command("add")
def supplier_add_command(
    name: str = typer.Option(..., "--name"),
    status: str = typer.Option("unknown", "--status"),
    comment: str | None = typer.Option(None, "--comment"),
) -> None:
    with SessionLocal() as session:
        row = SupplierService(session).add_supplier(name=name, status=status, comment=comment)
    typer.echo(f"Supplier saved: {row.name} ({row.status})")


@supplier_cli.command("list")
def supplier_list_command() -> None:
    with SessionLocal() as session:
        rows = SupplierService(session).list_suppliers()
    if not rows:
        typer.echo("No suppliers")
        return
    for row in rows:
        typer.echo(f"{row.name} | status={row.status} | rating={row.rating if row.rating is not None else '-'}")


@supplier_cli.command("update")
def supplier_update_command(
    name: str = typer.Option(..., "--name"),
    status: str = typer.Option(..., "--status"),
    comment: str | None = typer.Option(None, "--comment"),
) -> None:
    with SessionLocal() as session:
        row = SupplierService(session).update_supplier(name=name, status=status, comment=comment)
    typer.echo(f"Supplier updated: {row.name} ({row.status})")


@strategy_cli.command("create-defaults")
def strategy_create_defaults_command() -> None:
    with SessionLocal() as session:
        created = create_default_strategies(session)
    typer.echo(f"Default strategies created: {created}")


@strategy_cli.command("list")
def strategy_list_command() -> None:
    with SessionLocal() as session:
        rows = list_strategies(session)
    if not rows:
        typer.echo("No strategies")
        return
    for row in rows:
        typer.echo(
            f"{row.name} | active={row.is_active} | min_margin={row.min_margin_percent} | "
            f"min_profit={row.min_profit_amount} | max_risk={row.max_risk_level}"
        )


@strategy_cli.command("activate")
def strategy_activate_command(name: str = typer.Option(..., "--name")) -> None:
    with SessionLocal() as session:
        row = activate_strategy(session, name=name)
    typer.echo(f"Active strategy: {row.name}")


@strategy_cli.command("show-active")
def strategy_show_active_command() -> None:
    with SessionLocal() as session:
        row = get_active_strategy(session)
    if row is None:
        typer.echo("No active strategy")
        return
    typer.echo(
        f"{row.name} | min_margin={row.min_margin_percent} | min_profit={row.min_profit_amount} | "
        f"max_risk={row.max_risk_level}"
    )


@watchlist_cli.command("add")
def watchlist_add_command(
    purchase_id: int = typer.Option(..., "--purchase-id"),
    note: str | None = typer.Option(None, "--note"),
) -> None:
    with SessionLocal() as session:
        row = WatchlistService(session).add(purchase_id=purchase_id, note=note, status="watch")
    typer.echo(f"Watchlist updated: purchase_id={row.purchase_id} status={row.status}")


@watchlist_cli.command("list")
def watchlist_list_command() -> None:
    with SessionLocal() as session:
        rows = WatchlistService(session).list()
    if not rows:
        typer.echo("Watchlist empty")
        return
    for row in rows:
        typer.echo(f"purchase_id={row.purchase_id} status={row.status} note={row.note or '-'}")


@watchlist_cli.command("update")
def watchlist_update_command(
    purchase_id: int = typer.Option(..., "--purchase-id"),
    status: str = typer.Option(..., "--status"),
    note: str | None = typer.Option(None, "--note"),
) -> None:
    with SessionLocal() as session:
        row = WatchlistService(session).update(purchase_id=purchase_id, status=status, note=note)
    typer.echo(f"Watchlist updated: purchase_id={row.purchase_id} status={row.status}")


@watchlist_cli.command("remove")
def watchlist_remove_command(purchase_id: int = typer.Option(..., "--purchase-id")) -> None:
    with SessionLocal() as session:
        removed = WatchlistService(session).remove(purchase_id=purchase_id)
    typer.echo("Removed" if removed else "Not found")


@user_cli.command("create")
def user_create_command(
    username: str = typer.Option(..., "--username"),
    email: str = typer.Option(..., "--email"),
    role: str = typer.Option("viewer", "--role"),
    password: str | None = typer.Option(None, "--password"),
) -> None:
    with SessionLocal() as session:
        row = UserService(session).create_user(username=username, email=email, role=role, password=password)
    typer.echo(f"User created: {row.username} ({row.role})")


@user_cli.command("set-password")
def user_set_password_command(
    username: str = typer.Option(..., "--username"),
    password: str = typer.Option(..., "--password"),
) -> None:
    with SessionLocal() as session:
        row = UserService(session).set_password(username=username, password=password)
    typer.echo(f"Password updated for {row.username}")


@user_cli.command("list")
def user_list_command() -> None:
    with SessionLocal() as session:
        rows = UserService(session).list_users()
    if not rows:
        typer.echo("No users")
        return
    for row in rows:
        typer.echo(f"{row.username} | {row.email} | role={row.role} | active={row.is_active}")


@user_cli.command("disable")
def user_disable_command(username: str = typer.Option(..., "--username")) -> None:
    with SessionLocal() as session:
        row = UserService(session).set_active(username=username, is_active=False)
    typer.echo(f"User disabled: {row.username}")


@user_cli.command("enable")
def user_enable_command(username: str = typer.Option(..., "--username")) -> None:
    with SessionLocal() as session:
        row = UserService(session).set_active(username=username, is_active=True)
    typer.echo(f"User enabled: {row.username}")


@demo_cli.command("seed")
def seed_demo_command() -> None:
    with SessionLocal() as session:
        result = seed_demo_data(session)
    typer.echo(f"Demo data seeded: purchases={result['purchases']} offers={result['offers']}")


@demo_cli.command("reset")
def reset_demo_command(force_demo: bool = typer.Option(False, "--force-demo")) -> None:
    with SessionLocal() as session:
        reset_demo_data(session, force_demo=force_demo)
    typer.echo("Demo data reset complete")


@cli.command("seed-demo")
def seed_demo_alias() -> None:
    seed_demo_command()


@cli.command("reset-demo")
def reset_demo_alias(force_demo: bool = typer.Option(False, "--force-demo")) -> None:
    reset_demo_command(force_demo=force_demo)


@db_cli.command("current")
def db_current_command() -> None:
    settings = get_settings()
    backend = database_backend(settings.database_url)
    mode = migration_mode_for_settings(settings)
    typer.echo(f"database_backend={backend}")
    typer.echo(f"migrations_mode={mode}")
    if backend == "sqlite" and settings.app_mode == "production":
        typer.echo("SQLite is not supported in production. Use PostgreSQL.")
        raise typer.Exit(code=1)
    _check_database_connection_or_exit()
    if mode == SQLITE_DEMO_MIGRATION_MODE:
        typer.echo("current_revision=not_applicable_for_sqlite_demo")
        typer.echo("head_revision=not_applicable_for_sqlite_demo")
        return
    try:
        typer.echo(f"current_revision={get_current_revision()}")
        typer.echo(f"head_revision={get_head_revision()}")
    except Exception as exc:  # noqa: BLE001
        typer.echo(f"Could not read Alembic revisions: {exc}")
        raise typer.Exit(code=1)


@db_cli.command("upgrade")
def db_upgrade_command() -> None:
    settings = get_settings()
    backend = database_backend(settings.database_url)
    mode = migration_mode_for_settings(settings)
    if backend == "sqlite" and settings.app_mode == "production":
        typer.echo("SQLite is not supported in production. Use PostgreSQL.")
        raise typer.Exit(code=1)
    _check_database_connection_or_exit()
    if mode == SQLITE_DEMO_MIGRATION_MODE:
        init_database_for_settings(settings)
        typer.echo("SQLite demo schema ensured via metadata.create_all")
        return
    upgrade_head()
    typer.echo("Database upgraded to head")


@db_cli.command("check")
def db_check_command() -> None:
    settings = get_settings()
    backend = database_backend(settings.database_url)
    mode = migration_mode_for_settings(settings)
    typer.echo(f"database_backend={backend}")
    typer.echo(f"migrations_mode={mode}")
    if backend == "sqlite" and settings.app_mode == "production":
        typer.echo("SQLite is not supported in production. Use PostgreSQL.")
        raise typer.Exit(code=1)

    _check_database_connection_or_exit()

    if mode == SQLITE_DEMO_MIGRATION_MODE:
        typer.echo("migrations_up_to_date=True (sqlite demo uses metadata.create_all)")
        return
    try:
        current_revision = get_current_revision()
        head_revision = get_head_revision()
    except Exception as exc:  # noqa: BLE001
        typer.echo(f"Could not verify Alembic revisions: {exc}")
        raise typer.Exit(code=1)

    typer.echo(f"current_revision={current_revision}")
    typer.echo(f"head_revision={head_revision}")
    up_to_date = current_revision == head_revision
    typer.echo(f"migrations_up_to_date={up_to_date}")
    if not up_to_date:
        raise typer.Exit(code=1)


@cli.command("smoke-test")
def smoke_test_command() -> None:
    settings = get_settings()
    ensure_runtime_directories(settings)
    smoke_log = settings.logs_dir / "smoke-test.log"
    lines: list[str] = [f"smoke-test started at {utc_now().isoformat()}"]
    failures: list[str] = []

    def run_step(name: str, fn, required: bool = True) -> None:
        try:
            fn()
            lines.append(f"[PASS] {name}")
            typer.echo(f"[PASS] {name}")
        except Exception as exc:  # noqa: BLE001
            prefix = "[FAIL]" if required else "[SKIP]"
            lines.append(f"{prefix} {name}: {exc}")
            typer.echo(f"{prefix} {name}: {exc}")
            if required:
                failures.append(f"{name}: {exc}")

    if settings.real_network_enabled:
        lines.append("[WARN] REAL_NETWORK_ENABLED=true, smoke-test still uses stub/demo path")
        typer.echo("smoke-test warning: REAL_NETWORK_ENABLED=true, using stub/demo path only")

    run_step("init-db", init_db, required=True)
    run_step("seed-demo", seed_demo_command, required=True)
    run_step("extract-attributes --all", lambda: extract_attributes_command(all=True, purchase_id=None, item_id=None), required=False)
    run_step("search-prices --mode stub", lambda: search_prices_command(mode="stub", limit=10, purchase_id=None, item_id=None), required=True)
    run_step("validate-data", lambda: validate_data_command(purchase_id=None), required=True)
    run_step("calculate", calculate_command, required=True)
    run_step("financial-check", lambda: financial_check_command(purchase_id=None), required=True)
    run_step("evaluate", lambda: evaluate_command(purchase_id=None), required=True)
    run_step("export-excel", lambda: export_excel_command(output_path=Path("exports/smoke_export.xlsx")), required=True)
    run_step("export-decisions", lambda: export_decisions_command(file=Path("exports/smoke_decisions.xlsx")), required=False)
    run_step("healthcheck", healthcheck_command, required=True)
    run_step("doctor", doctor_command, required=True)

    lines.append(f"smoke-test finished at {utc_now().isoformat()}")
    smoke_log.write_text("\n".join(lines) + "\n", encoding="utf-8")
    typer.echo(f"smoke log: {smoke_log}")

    if failures:
        raise typer.Exit(code=1)
    typer.echo("smoke-test completed")


@parse_cli.command("run")
def parse_group_run_command(
    source: str = typer.Option("mos_portal", "--source"),
    limit: int | None = typer.Option(None, "--limit"),
    status: str = typer.Option("Прием предложений", "--status"),
) -> None:
    parse_command(source=source, limit=limit, dry_run=False, save_raw=True, status=status)


@price_search_cli.command("run")
def price_search_group_run_command(
    mode: str = typer.Option("manual", "--mode"),
    limit: int | None = typer.Option(None, "--limit"),
) -> None:
    search_prices_command(mode=mode, limit=limit, purchase_id=None, item_id=None)


@diagnostics_cli.command("doctor")
def diagnostics_doctor_command() -> None:
    doctor_command()


@diagnostics_cli.command("health")
def diagnostics_health_command() -> None:
    healthcheck_command()


@scheduler_cli.command("run")
def scheduler_run_group_command() -> None:
    run_scheduler_command()


@scheduler_cli.command("status")
def scheduler_status_group_command() -> None:
    scheduler_status_command()




def _check_database_connection_or_exit() -> None:
    with SessionLocal() as session:
        try:
            session.execute(text("SELECT 1"))
            typer.echo("database connection: ok")
        except Exception as exc:  # noqa: BLE001
            typer.echo(f"database connection failed: {exc}")
            raise typer.Exit(code=1)


def _required_hosts_for_source(source: str) -> list[str]:
    if source == "mos_portal":
        return ["zakupki.mos.ru", ".zakupki.mos.ru", "api.zakupki.mos.ru"]
    return ["agregatoreat.ru", ".agregatoreat.ru"]


def _run_real_source_probe(source: str, status: str, limit: int) -> dict[str, Any]:
    warnings: list[str] = []
    errors: list[str] = []
    parsed: list[Any] = []
    api_ok = False
    browser_fallback_used = False
    raw_records: list[dict[str, Any]] = []

    if source == "mos_portal":
        api_client = MosPortalApiClient()
        raw_records, api_warnings, api_errors = api_client.fetch_purchases(status=status, limit=limit)
        warnings.extend(api_warnings)
        errors.extend(api_errors)
        if raw_records:
            api_ok = True
        else:
            fallback_records, fb_warnings, fb_errors = MosPortalBrowserFallback().fetch_cards(status=status, limit=limit)
            browser_fallback_used = True
            warnings.extend(fb_warnings)
            errors.extend(fb_errors)
            raw_records = fallback_records

        for raw in raw_records[:limit]:
            try:
                enriched = api_client.enrich_with_details(raw)
                parsed.append(parse_mos_purchase_payload(enriched, source="mos_portal"))
            except Exception as exc:  # noqa: BLE001
                errors.append(f"parse error: {exc}")
    else:
        api_client = EatApiClient()
        raw_records, api_warnings, api_errors = api_client.fetch_purchases(status=status, limit=limit)
        warnings.extend(api_warnings)
        errors.extend(api_errors)
        if raw_records:
            api_ok = True
        else:
            fallback_records, fb_warnings, fb_errors = EatBrowserFallback().fetch_cards(status=status, limit=limit)
            warnings.extend(fb_warnings)
            errors.extend(fb_errors)
            raw_records = fallback_records
            browser_fallback_used = bool(raw_records)

        for raw in raw_records[:limit]:
            try:
                parsed.append(parse_eat_purchase_payload(raw, source="eat"))
            except Exception as exc:  # noqa: BLE001
                errors.append(f"parse error: {exc}")

    lowered = " ".join([*warnings, *errors]).lower()
    auth_required = any(mark in lowered for mark in ["401", "403", "unauthorized", "auth"])
    captcha_or_blocked = any(mark in lowered for mark in ["captcha", "blocked", "challenge", "timeout"])

    connector_status = "failed"
    if api_ok:
        connector_status = "api_ok"
    elif browser_fallback_used:
        connector_status = "browser_fallback_used"
    elif auth_required:
        connector_status = "auth_required"
    elif captcha_or_blocked:
        connector_status = "captcha_or_blocked"

    return {
        "api_ok": api_ok,
        "browser_fallback_used": browser_fallback_used,
        "auth_required": auth_required,
        "captcha_or_blocked": captcha_or_blocked,
        "connector_status": connector_status,
        "warnings": warnings,
        "errors": errors,
        "parsed": parsed,
        "raw_records_count": len(raw_records),
    }


def _print_parsed_sample(parsed: list[Any], limit: int) -> None:
    if not parsed:
        typer.echo("No parsed purchases")
        return
    for idx, purchase in enumerate(parsed[:limit], start=1):
        items_count = len(getattr(purchase, "items", []) or [])
        missing_fields = []
        for name in ("title", "url", "status", "region", "submission_deadline", "max_total_price"):
            value = getattr(purchase, name, None)
            if value in (None, "", []):
                missing_fields.append(name)
        typer.echo(
            f"[{idx}] external_id={purchase.external_id} title={purchase.title!r} "
            f"items={items_count} missing={','.join(missing_fields) if missing_fields else '-'}"
        )


def _parsed_validation_metrics(parsed: list[Any]) -> dict[str, Any]:
    missing: dict[str, int] = {}
    valid = 0
    for purchase in parsed:
        local_missing: list[str] = []
        if not getattr(purchase, "external_id", None):
            local_missing.append("external_id")
        deadline = getattr(purchase, "submission_deadline", None)
        if deadline is None:
            local_missing.append("submission_deadline")
        price = getattr(purchase, "max_total_price", None)
        if price is None or float(price) <= 0:
            local_missing.append("max_total_price")

        for item in getattr(purchase, "items", []) or []:
            quantity = getattr(item, "quantity", None)
            if quantity is None or float(quantity) <= 0:
                local_missing.append("quantity")
            max_unit_price = getattr(item, "max_unit_price", None)
            if max_unit_price is None:
                local_missing.append("max_unit_price")

        if not local_missing:
            valid += 1
        for key in local_missing:
            missing[key] = missing.get(key, 0) + 1
    return {"found": len(parsed), "valid": valid, "missing_fields": missing}


def _env_to_dict(lines: list[str]) -> dict[str, str]:
    values: dict[str, str] = {}
    for line in lines:
        if not line or line.lstrip().startswith("#") or "=" not in line:
            continue
        key, value = line.split("=", 1)
        values[key.strip()] = value.strip()
    return values


def _render_env_lines(lines: list[str], values: dict[str, str]) -> list[str]:
    rendered: list[str] = []
    for line in lines:
        if not line or line.lstrip().startswith("#") or "=" not in line:
            rendered.append(line)
            continue
        key, _value = line.split("=", 1)
        key = key.strip()
        rendered.append(f"{key}={values.get(key, '')}")
    return rendered

def _is_port_open(host: str, port: int) -> bool:
    with socket.socket(socket.AF_INET, socket.SOCK_STREAM) as sock:
        sock.settimeout(0.5)
        return sock.connect_ex((host, port)) == 0



def _safe_obj(value):
    if value is None:
        return None
    if hasattr(value, "__dict__"):
        return value.__dict__
    return str(value)



def json_dumps(payload: dict) -> str:
    return json.dumps(payload, ensure_ascii=False)


if __name__ == "__main__":
    cli()
