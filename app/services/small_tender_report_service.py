from __future__ import annotations

import csv
import io
import json
import re
import subprocess
import zipfile
import xml.etree.ElementTree as ET
from urllib.parse import urlparse

from dataclasses import dataclass
from pathlib import Path
from typing import Iterable

from openpyxl import Workbook
import requests


@dataclass
class AttachmentExtractionResult:
    purchase_external_id: str
    source_path: str
    status: str
    reason: str | None
    text: str


def generate_small_tender_report(
    *,
    market_csv: Path,
    tender_ref_csv: Path,
    out_csv: Path,
    out_xlsx: Path,
    diagnostics_csv: Path,
    attachments_manifest_csv: Path | None = None,
) -> dict[str, int]:
    market_rows = _read_csv(market_csv)
    ref_rows = _read_csv(tender_ref_csv)
    vendor_candidate_by_purchase = _build_vendor_candidate_index(market_rows)

    ref_price_by_purchase: dict[str, float] = {}
    for row in ref_rows:
        pid = str(row.get("purchase_external_id", "")).strip()
        price = _pick_first_price(
            row,
            [
                "tender_unit_price_ref",
                "unit_price",
                "offered_unit_price",
                "market_price",
                "min_price",
                "final_price",
                "price",
            ],
        )
        if not pid or price is None:
            continue
        prev = ref_price_by_purchase.get(pid)
        if prev is None or price < prev:
            ref_price_by_purchase[pid] = price

    attachments_by_purchase = _load_attachments_manifest(attachments_manifest_csv)
    extraction_by_purchase: dict[str, AttachmentExtractionResult] = {}
    for purchase_id, paths in attachments_by_purchase.items():
        extraction_by_purchase[purchase_id] = _extract_first_successful_attachment(purchase_id, paths)
    mos_portal_endpoint_cache: dict[str, AttachmentExtractionResult] = {}

    report_rows: list[dict[str, object]] = []
    diag_rows: list[dict[str, object]] = []
    input_rows_count = len(market_rows)
    excluded_non_goods_count = 0

    for row in market_rows:
        purchase_id = str(row.get("purchase_external_id", "")).strip()
        item_name = str(row.get("item_name", "")).strip()
        offer_title = str(row.get("offer_title", "")).strip()
        offer_source_url = str(row.get("offer_source_url", "")).strip()
        found_offer_unit_price = _pick_first_price(
            row,
            [
                "found_offer_unit_price",
                "offered_unit_price",
                "unit_price",
                "effective_unit_price",
                "market_price",
                "min_price",
                "final_price",
                "price",
            ],
        )
        market_unit_price = _pick_first_price(
            row,
            [
                "market_unit_price",
                "unit_price",
                "offered_unit_price",
                "market_price",
                "found_offer_unit_price",
                "effective_unit_price",
                "min_price",
                "final_price",
                "price",
            ],
        )
        market_price_source_note = "from_input"
        market_price_source_domain = _extract_domain(offer_source_url)
        market_source_domain_blocked = False
        has_valid_url = bool(offer_source_url and offer_source_url.startswith(("http://", "https://")))
        is_procurement_domain = _is_invalid_market_source_domain(market_price_source_domain)
        is_mos_portal = _is_mos_portal_row(row)
        has_vendor_candidate_for_purchase = vendor_candidate_by_purchase.get(purchase_id, False)

        if is_mos_portal and (market_unit_price is None or found_offer_unit_price is None):
            fallback_mos_price = _pick_first_price(
                row,
                [
                    "market_unit_price",
                    "found_offer_unit_price",
                    "effective_unit_price",
                    "offered_unit_price",
                    "unit_price",
                    "market_price",
                    "min_price",
                    "final_price",
                    "price",
                    "tender_unit_price_ref",
                ],
            )
            if fallback_mos_price is not None:
                if market_unit_price is None:
                    market_unit_price = fallback_mos_price
                if found_offer_unit_price is None:
                    found_offer_unit_price = fallback_mos_price
                market_price_source_note = "from_input_mos_portal_fallback"

        if _is_non_goods_item(item_name):
            excluded_non_goods_count += 1
            diag_rows.append(
                {
                    "purchase_external_id": purchase_id,
                    "attachment_path": "",
                    "tz_extraction_status": "skipped",
                    "tz_extraction_reason": "excluded_non_goods",
                    "tz_text_len": 0,
                    "tz_match_type": "skipped",
                    "strict_full_match": False,
                    "tz_overlap_ratio": 0.0,
                    "market_price_source_url": offer_source_url,
                    "market_price_source_domain": _extract_domain(offer_source_url),
                    "market_price_source_note": "from_input",
                    "margin_pct": None,
                    "decision_status": "excluded",
                    "decision_reason": "excluded_non_goods",
                }
            )
            continue

        if is_procurement_domain:
            original_market_unit_price = market_unit_price
            market_unit_price = None
            if (
                found_offer_unit_price is not None
                and original_market_unit_price is not None
                and found_offer_unit_price == original_market_unit_price
            ):
                found_offer_unit_price = None
            market_source_domain_blocked = True
            if has_vendor_candidate_for_purchase:
                market_price_source_note = "procurement_domain_blocked/vendor_candidate_preferred"
            else:
                market_price_source_note = "invalid_market_source_domain/procurement_domain_blocked"

        if is_mos_portal and not market_source_domain_blocked and (market_unit_price is None or found_offer_unit_price is None):
            fallback_mos_price = _pick_first_price(
                row,
                [
                    "market_unit_price",
                    "found_offer_unit_price",
                    "effective_unit_price",
                    "offered_unit_price",
                    "unit_price",
                    "market_price",
                    "min_price",
                    "final_price",
                    "price",
                    "tender_unit_price_ref",
                ],
            )
            if fallback_mos_price is not None:
                if market_unit_price is None:
                    market_unit_price = fallback_mos_price
                if found_offer_unit_price is None:
                    found_offer_unit_price = fallback_mos_price
                market_price_source_note = "from_input_mos_portal_fallback"
        extraction = extraction_by_purchase.get(purchase_id)
        if extraction is None or extraction.status != "ok":
            if _is_mos_portal_row(row):
                endpoint_extraction = mos_portal_endpoint_cache.get(purchase_id)
                if endpoint_extraction is None:
                    endpoint_extraction = _extract_tz_from_mos_portal_endpoint(purchase_id)
                    mos_portal_endpoint_cache[purchase_id] = endpoint_extraction
                if endpoint_extraction.status == "ok":
                    extraction = endpoint_extraction
                elif extraction is None:
                    extraction = endpoint_extraction
        if extraction is None:
            extraction = AttachmentExtractionResult(
                purchase_external_id=purchase_id,
                source_path="",
                status="failed",
                reason="no_attachment_manifest_entry",
                text="",
            )

        tz_text = extraction.text
        tz_tokens = _tokenize(tz_text)
        item_tokens = _tokenize(item_name)
        title_tokens = _tokenize(offer_title)
        all_offer_tokens = item_tokens | title_tokens

        if is_mos_portal:
            strict_full_match = _is_mos_portal_strict_attribute_match(
                tz_text=tz_text,
                item_name=item_name,
                offer_title=offer_title,
            )
        else:
            strict_full_match = _is_strict_attribute_match(tz_text=tz_text, offer_text=f"{item_name} {offer_title}") or (
                bool(tz_tokens) and tz_tokens.issubset(all_offer_tokens)
            )
        overlap = (len(tz_tokens & all_offer_tokens) / len(tz_tokens)) if tz_tokens else 0.0
        fallback_non_product_match = False
        if not strict_full_match and _is_non_product_tz_text(tz_text):
            fallback_non_product_match = bool(item_tokens or title_tokens)

        has_valid_search_price = market_unit_price is not None and found_offer_unit_price is not None
        full_match_vendor_eligible = (
            (strict_full_match or fallback_non_product_match)
            and not is_procurement_domain
            and has_valid_search_price
            and (has_valid_url or not is_mos_portal)
        )

        match_type = "none"
        if full_match_vendor_eligible:
            match_type = "full"
        elif strict_full_match or fallback_non_product_match or overlap >= 0.5:
            match_type = "partial"

        tender_ref_price = ref_price_by_purchase.get(purchase_id)
        margin_pct = None
        if market_unit_price is not None and tender_ref_price and tender_ref_price > 0:
            margin_pct = round(((tender_ref_price - market_unit_price) / tender_ref_price) * 100, 2)

        risk_level = "unknown"
        decision_status = "reject"
        decision_reason = ""

        if extraction.status != "ok":
            decision_status = "reject"
            decision_reason = f"tz_extraction_failed:{extraction.reason or 'unknown'}"
            risk_level = "critical"
        elif match_type != "full":
            decision_status = "reject"
            if is_procurement_domain:
                decision_reason = "procurement_domain_blocked"
            elif is_mos_portal and not has_valid_url:
                decision_reason = "missing_or_invalid_offer_source_url"
            elif not has_valid_search_price:
                decision_reason = "missing_search_price"
            else:
                decision_reason = f"strict_full_match_required:{match_type}"
            risk_level = "critical"
        elif market_unit_price is None or found_offer_unit_price is None:
            decision_status = "reject"
            decision_reason = "missing_search_price"
            risk_level = "critical"
        elif margin_pct is None:
            decision_status = "reject"
            decision_reason = "missing_price_for_margin"
            risk_level = "critical"
        elif margin_pct > 50:
            decision_status = "reject"
            decision_reason = "margin_gt_50_reject"
            risk_level = "critical"
        elif margin_pct > 25:
            decision_status = "high_risk"
            decision_reason = "margin_gt_25_high_risk"
            risk_level = "high"
        else:
            decision_status = "green"
            decision_reason = "strict_full_match_and_margin_ok"
            risk_level = "low"

        report_rows.append(
            {
                **row,
                "tz_extraction_status": extraction.status,
                "tz_extraction_reason": extraction.reason or "",
                "tz_attachment_path": extraction.source_path,
                "tz_text_len": len(tz_text),
                "tz_match_type": match_type,
                "tz_overlap_ratio": round(overlap, 4),
                "strict_full_match": full_match_vendor_eligible,
                "tender_unit_price_ref": tender_ref_price,
                "market_unit_price": market_unit_price,
                "found_offer_unit_price": found_offer_unit_price,
                "offer_source_url": offer_source_url,
                "market_price_source_url": offer_source_url,
                "market_price_source_domain": market_price_source_domain,
                "market_price_source_note": market_price_source_note,
                "source_selection_reason": _source_selection_reason(
                    is_procurement_domain=is_procurement_domain,
                    has_vendor_candidate_for_purchase=has_vendor_candidate_for_purchase,
                    has_valid_url=has_valid_url,
                    has_valid_search_price=has_valid_search_price,
                    strict_full_match=full_match_vendor_eligible,
                ),
                "margin_pct": margin_pct,
                "decision_status": decision_status,
                "risk_level": risk_level,
                "decision_reason": decision_reason,
            }
        )

        diag_rows.append(
            {
                "purchase_external_id": purchase_id,
                "attachment_path": extraction.source_path,
                "tz_extraction_status": extraction.status,
                "tz_extraction_reason": extraction.reason or "",
                "tz_text_len": len(tz_text),
                "tz_match_type": match_type,
                "strict_full_match": full_match_vendor_eligible,
                "tz_overlap_ratio": round(overlap, 4),
                "market_price_source_url": offer_source_url,
                "market_price_source_domain": market_price_source_domain,
                "market_price_source_note": market_price_source_note,
                "source_selection_reason": _source_selection_reason(
                    is_procurement_domain=is_procurement_domain,
                    has_vendor_candidate_for_purchase=has_vendor_candidate_for_purchase,
                    has_valid_url=has_valid_url,
                    has_valid_search_price=has_valid_search_price,
                    strict_full_match=full_match_vendor_eligible,
                ),
                "margin_pct": margin_pct,
                "decision_status": decision_status,
                "decision_reason": decision_reason,
            }
        )

    _write_csv(out_csv, report_rows)
    _write_xlsx(out_xlsx, report_rows)
    _write_csv(diagnostics_csv, diag_rows)

    return {
        "input_rows": input_rows_count,
        "report_rows": len(report_rows),
        "excluded_non_goods_rows": excluded_non_goods_count,
        "diagnostics_rows": len(diag_rows),
        "margin_0_25_rows": sum(
            1 for r in report_rows if r.get("margin_pct") is not None and 0 <= float(r["margin_pct"]) <= 25
        ),
        "margin_gt_25_rows": sum(1 for r in report_rows if r.get("margin_pct") is not None and float(r["margin_pct"]) > 25),
        "margin_gt_50_rows": sum(1 for r in report_rows if r.get("margin_pct") is not None and float(r["margin_pct"]) > 50),
        "full_match_rows": sum(1 for r in report_rows if r.get("tz_match_type") == "full"),
        "green_rows": sum(1 for r in report_rows if r.get("decision_status") == "green"),
        "high_risk_rows": sum(1 for r in report_rows if r.get("decision_status") == "high_risk"),
        "reject_rows": sum(1 for r in report_rows if r.get("decision_status") == "reject"),
    }


def _load_attachments_manifest(path: Path | None) -> dict[str, list[Path]]:
    if path is None or not path.exists():
        return {}

    rows = _read_csv(path)
    result: dict[str, list[Path]] = {}
    for row in rows:
        purchase_id = str(row.get("purchase_external_id", "")).strip()
        attachment_path = str(row.get("attachment_path", "")).strip()
        if not purchase_id or not attachment_path:
            continue
        result.setdefault(purchase_id, []).append(Path(attachment_path))
    return result


def _extract_first_successful_attachment(purchase_id: str, paths: Iterable[Path]) -> AttachmentExtractionResult:
    errors: list[str] = []
    first_path = ""
    for path in paths:
        first_path = str(path)
        if not path.exists():
            errors.append(f"not_found:{path}")
            continue
        try:
            text = extract_text_from_attachment(path)
            cleaned = _clean_text(text)
            if cleaned:
                return AttachmentExtractionResult(
                    purchase_external_id=purchase_id,
                    source_path=str(path),
                    status="ok",
                    reason=None,
                    text=cleaned,
                )
            errors.append(f"empty_text:{path}")
        except Exception as exc:  # pragma: no cover - defensive
            errors.append(f"extract_error:{path}:{exc}")

    reason = ";".join(errors) if errors else "no_attachments"
    return AttachmentExtractionResult(
        purchase_external_id=purchase_id,
        source_path=first_path,
        status="failed",
        reason=reason,
        text="",
    )


def _is_mos_portal_row(row: dict[str, object]) -> bool:
    source = str(row.get("source", "")).strip().lower()
    if source == "mos_portal":
        return True
    source_name = str(row.get("source_name", "")).strip().lower()
    if source_name == "mos_portal":
        return True
    url = str(row.get("purchase_url", "") or row.get("offer_source_url", "")).strip().lower()
    return "zakupki.mos.ru" in url


def _extract_tz_from_mos_portal_endpoint(purchase_id: str) -> AttachmentExtractionResult:
    if not purchase_id:
        return AttachmentExtractionResult(
            purchase_external_id=purchase_id,
            source_path="",
            status="failed",
            reason="empty_purchase_external_id",
            text="",
        )

    session = requests.Session()
    session.trust_env = False
    session.headers.update({"User-Agent": "Mozilla/5.0", "Accept": "*/*"})

    endpoint_payloads = _mos_portal_fetch_payloads(session, purchase_id)
    files = _mos_portal_files_from_payloads(endpoint_payloads)
    if not files:
        payload_text = _mos_portal_payload_tz_text(endpoint_payloads)
        if payload_text:
            return AttachmentExtractionResult(
                purchase_external_id=purchase_id,
                source_path="mos_portal_api_payload",
                status="ok",
                reason=None,
                text=payload_text,
            )
        return AttachmentExtractionResult(
            purchase_external_id=purchase_id,
            source_path="",
            status="failed",
            reason="auction_get_no_files",
            text="",
        )

    errors: list[str] = []
    for file_id, file_name in files:
        if not _is_likely_tz_name(file_name):
            errors.append(f"rejected_non_tz_name:{file_id}:{file_name}")
            continue
        try:
            data, content_type = _mos_portal_download_file(session, file_id)
        except Exception as exc:  # pragma: no cover - network/runtime branch
            errors.append(f"download_failed:{file_id}:{exc}")
            continue
        if len(data) < 512:
            errors.append(f"too_small:{file_id}:{len(data)}")
            continue

        text = _clean_text(_extract_text_from_bytes(file_name, data, content_type))
        if not text:
            errors.append(f"rejected_empty_text:{file_id}:{file_name}")
            continue

        return AttachmentExtractionResult(
            purchase_external_id=purchase_id,
            source_path=f"mos_portal_api:{file_id}:{file_name}",
            status="ok",
            reason=None,
            text=text,
        )

    return AttachmentExtractionResult(
        purchase_external_id=purchase_id,
        source_path="",
        status="failed",
        reason=";".join(errors) if errors else "all_candidates_rejected",
        text="",
    )


def _is_likely_tz_name(name: str) -> bool:
    lower = name.lower().replace("ё", "е")
    if re.search(r"(регистрац|эп|инструкц|памятк|guide|widget|attach\\.svg)", lower):
        return False
    return bool(re.search(r"(тз|техническ|задани|spec|специф|описани)", lower))


def _mos_portal_fetch_payloads(session: requests.Session, auction_id: str) -> list[tuple[str, dict[str, object]]]:
    endpoints: tuple[tuple[str, str], ...] = (
        ("https://zakupki.mos.ru/newapi/api/Auction/Get", "auctionId"),
        ("https://zakupki.mos.ru/newapi/api/Purchase/Get", "purchaseId"),
        ("https://zakupki.mos.ru/newapi/api/Need/Get", "needId"),
    )
    payloads: list[tuple[str, dict[str, object]]] = []

    for url, param_name in endpoints:
        try:
            response = session.get(url, params={param_name: auction_id}, timeout=60)
        except Exception:
            continue
        if response.status_code != 200:
            continue
        try:
            payload = response.json()
        except Exception:
            continue
        payloads.append((url, payload))

    return payloads


def _mos_portal_files_from_payloads(payloads: list[tuple[str, dict[str, object]]]) -> list[tuple[int, str]]:
    for _, payload in payloads:
        files = payload.get("files") or []
        items: list[tuple[int, str]] = []
        seen: set[int] = set()
        for entry in files:
            try:
                file_id = int(entry.get("id"))
            except Exception:
                continue
            if file_id in seen:
                continue
            seen.add(file_id)
            name = str(entry.get("name") or entry.get("fileName") or f"Download_{file_id}")
            items.append((file_id, name))

        if items:
            return sorted(items, key=lambda pair: (0 if _is_likely_tz_name(pair[1]) else 1, pair[1].lower()))

    return []


def _mos_portal_payload_tz_text(payloads: list[tuple[str, dict[str, object]]]) -> str:
    for _, payload in reversed(payloads):
        if not isinstance(payload, dict):
            continue
        parts: list[str] = []
        for key in ("name", "description", "itemDescription"):
            value = payload.get(key)
            if isinstance(value, str):
                parts.append(value)

        for entry in payload.get("items") or []:
            if not isinstance(entry, dict):
                continue
            for key in ("name", "description", "itemDescription"):
                value = entry.get(key)
                if isinstance(value, str):
                    parts.append(value)

        cleaned = _clean_text("\n".join(parts))
        if cleaned:
            return cleaned

    return ""


def _mos_portal_download_file(session: requests.Session, file_id: int) -> tuple[bytes, str]:
    url = "https://zakupki.mos.ru/newapi/api/FileStorage/Download"
    response = session.get(url, timeout=60, allow_redirects=True, params={"id": file_id})
    response.raise_for_status()
    content_type = (response.headers.get("content-type") or "").lower()
    if "text/html" in content_type:
        raise RuntimeError("html_instead_of_file")
    return response.content, content_type


def _extract_text_from_bytes(name: str, data: bytes, content_type: str) -> str:
    suffix = Path(name).suffix.lower()
    if suffix == ".txt" or "text/plain" in content_type:
        return data.decode("utf-8", errors="ignore")
    if suffix == ".docx":
        return _extract_docx_text_from_bytes(data)
    if suffix == ".pdf" or "pdf" in content_type:
        return _extract_pdf_text_from_bytes(data)
    return ""


def _extract_docx_text_from_bytes(data: bytes) -> str:
    try:
        with zipfile.ZipFile(io.BytesIO(data)) as zf:
            xml = zf.read("word/document.xml")
        root = ET.fromstring(xml)
    except Exception:
        return ""
    return "".join((node.text or "") for node in root.iter() if node.text)


def _extract_pdf_text_from_bytes(data: bytes) -> str:
    try:
        from pypdf import PdfReader
    except Exception:
        return ""
    try:
        reader = PdfReader(io.BytesIO(data))
        return "\n".join((page.extract_text() or "") for page in reader.pages)
    except Exception:
        return ""


def extract_text_from_attachment(path: Path) -> str:
    suffix = path.suffix.lower()
    if suffix == ".txt":
        return path.read_text(encoding="utf-8", errors="ignore")
    if suffix == ".docx":
        return _extract_docx_text(path)
    if suffix == ".pdf":
        return _extract_pdf_text(path)
    raise ValueError(f"unsupported_attachment_type:{suffix}")


def _extract_docx_text(path: Path) -> str:
    with zipfile.ZipFile(path) as zf:
        data = zf.read("word/document.xml").decode("utf-8", errors="ignore")
    text = re.sub(r"<[^>]+>", " ", data)
    return _clean_text(text)


def _extract_pdf_text(path: Path) -> str:
    pdftotext_bin = "pdftotext"
    check = subprocess.run(["which", pdftotext_bin], capture_output=True, text=True)
    if check.returncode != 0:
        raise RuntimeError("pdftotext_not_installed")

    completed = subprocess.run([pdftotext_bin, "-layout", str(path), "-"], capture_output=True, text=True)
    if completed.returncode != 0:
        stderr = (completed.stderr or "").strip()
        raise RuntimeError(f"pdftotext_failed:{stderr or 'unknown_error'}")
    return completed.stdout


def _clean_text(text: str) -> str:
    normalized = text.replace("\xa0", " ")
    normalized = re.sub(r"\s+", " ", normalized)
    return normalized.strip()


def _tokenize(text: str) -> set[str]:
    prepared = text.lower().replace("ё", "е")
    prepared = re.sub(r"[^a-zа-я0-9]+", " ", prepared, flags=re.IGNORECASE)
    return {token for token in prepared.split() if len(token) >= 4}


def _is_non_goods_item(item_name: str) -> bool:
    text = _normalize_ru_text(item_name)
    non_goods_markers = (
        "монтаж",
        "демонтаж",
        "ремонт",
        "обслуживание",
        "установка",
        "наладка",
        "проектирование",
        "услуги",
        "работы",
    )
    return any(marker in text for marker in non_goods_markers)


def _normalize_ru_text(text: str) -> str:
    prepared = text.lower().replace("ё", "е")
    prepared = prepared.replace("\xa0", " ")
    prepared = re.sub(r"[^a-zа-я0-9]+", " ", prepared, flags=re.IGNORECASE)
    prepared = re.sub(r"\s+", " ", prepared)
    return prepared.strip()


def _extract_model_tokens(text: str) -> set[str]:
    prepared = re.sub(r"[^a-zа-я0-9\-/]+", " ", text.lower().replace("ё", "е"), flags=re.IGNORECASE)
    candidates = prepared.split()
    return {
        token
        for token in candidates
        if len(token) >= 4 and any(ch.isalpha() for ch in token) and any(ch.isdigit() for ch in token)
    }


def _extract_numbers(text: str) -> set[str]:
    return set(re.findall(r"\d{2,}", text))


def _extract_units(text: str) -> set[str]:
    units_map = {
        "шт": "шт",
        "штук": "шт",
        "штука": "шт",
        "штуки": "шт",
        "ед": "шт",
        "pcs": "шт",
        "pc": "шт",
        "компл": "компл",
        "комплект": "компл",
        "комплекта": "компл",
    }
    tokens = _normalize_ru_text(text).split()
    return {units_map[t] for t in tokens if t in units_map}


def _is_strict_attribute_match(*, tz_text: str, offer_text: str) -> bool:
    tz_norm = _normalize_ru_text(tz_text)
    offer_norm = _normalize_ru_text(offer_text)
    if not tz_norm or not offer_norm:
        return False

    offer_compact = offer_text.lower().replace("ё", "е").replace(" ", "")
    tz_models = _extract_model_tokens(tz_text)
    if tz_models and not all(model in offer_compact for model in tz_models):
        return False

    tz_numbers = _extract_numbers(tz_norm)
    offer_numbers = _extract_numbers(offer_norm)
    long_tz_numbers = {n for n in tz_numbers if len(n) >= 5}
    if long_tz_numbers and not long_tz_numbers.issubset(offer_numbers):
        return False

    tz_units = _extract_units(tz_norm)
    offer_units = _extract_units(offer_norm)
    if tz_units and offer_units and not tz_units.issubset(offer_units):
        return False

    return bool(tz_models or long_tz_numbers or (tz_units and tz_numbers))


def _is_mos_portal_strict_attribute_match(*, tz_text: str, item_name: str, offer_title: str) -> bool:
    tz_norm = _normalize_ru_text(tz_text)
    offer_text = f"{item_name} {offer_title}".strip()
    offer_norm = _normalize_ru_text(offer_text)
    if not tz_norm or not offer_norm:
        return False

    tz_compact = re.sub(r"[^a-zа-я0-9]+", "", tz_norm)
    tz_numbers = _extract_numbers(tz_norm)

    offer_models = _extract_model_tokens(offer_text)
    normalized_offer_models = {re.sub(r"[^a-zа-я0-9]+", "", model) for model in offer_models}
    normalized_offer_models = {m for m in normalized_offer_models if m}
    matched_models = {model for model in normalized_offer_models if model in tz_compact}
    if normalized_offer_models and matched_models != normalized_offer_models:
        return False

    offer_long_numbers = {n for n in _extract_numbers(offer_norm) if len(n) >= 5}
    matched_long_numbers = offer_long_numbers & tz_numbers
    if offer_long_numbers and not matched_long_numbers:
        return False

    offer_brands = _extract_brand_tokens(offer_norm)
    if offer_brands and not any(brand in tz_norm for brand in offer_brands):
        return False

    return bool(normalized_offer_models or (offer_brands and matched_long_numbers))


def _extract_brand_tokens(text: str) -> set[str]:
    known_brands = {
        "huawei",
        "hewlett",
        "packard",
        "dell",
        "lenovo",
        "cisco",
        "mikrotik",
        "zyxel",
        "canon",
        "xerox",
        "kyocera",
        "brother",
        "epson",
        "ricoh",
        "samsung",
        "apple",
        "asus",
        "acer",
        "intel",
        "amd",
        "nvidia",
        "hp",
        "lg",
    }
    tokens = set(text.split())
    matched = {token for token in tokens if token in known_brands}
    if "tp" in tokens and "link" in tokens:
        matched.update({"tp", "link"})
    return matched


def _is_non_product_tz_text(text: str) -> bool:
    norm = _normalize_ru_text(text)
    if not norm:
        return False

    tokens = norm.split()
    if len(tokens) < 5:
        return False

    if _extract_model_tokens(norm):
        return False
    if any(len(n) >= 5 for n in _extract_numbers(norm)):
        return False

    noise_markers = {
        "инструкция",
        "регистрация",
        "пользователя",
        "организации",
        "электронной",
        "подписи",
        "требования",
        "порядок",
        "документооборот",
    }
    overlap = sum(1 for t in tokens if t in noise_markers)
    return overlap >= 2


def _extract_domain(url: str) -> str:
    if not url:
        return ""
    try:
        return (urlparse(url).netloc or "").lower()
    except Exception:
        return ""


def _is_invalid_market_source_domain(domain: str) -> bool:
    normalized = (domain or "").strip().lower()
    if not normalized:
        return False

    blocked_domains = (
        "zakupki.mos.ru",
        "market.mosreg.ru",
        "business.roseltorg.ru",
        "roseltorg.ru",
        "rts-tender.ru",
        "sberbank-ast.ru",
        "etp-ets.ru",
        "tektorg.ru",
        "goszakupki.gov.ru",
        "zakupki.gov.ru",
        "zakupki360.ru",
        "zakupki360.com",
    )
    return any(normalized == d or normalized.endswith(f".{d}") for d in blocked_domains)


def _build_vendor_candidate_index(rows: list[dict[str, str]]) -> dict[str, bool]:
    result: dict[str, bool] = {}
    for row in rows:
        purchase_id = str(row.get("purchase_external_id", "")).strip()
        if not purchase_id:
            continue
        url = str(row.get("offer_source_url", "")).strip()
        domain = _extract_domain(url)
        has_valid_url = bool(url and url.startswith(("http://", "https://")))
        price = _pick_first_price(
            row,
            [
                "market_unit_price",
                "found_offer_unit_price",
                "offered_unit_price",
                "unit_price",
                "effective_unit_price",
                "market_price",
                "min_price",
                "final_price",
                "price",
            ],
        )
        is_vendor_candidate = has_valid_url and not _is_invalid_market_source_domain(domain) and price is not None
        if is_vendor_candidate:
            result[purchase_id] = True
        else:
            result.setdefault(purchase_id, False)
    return result


def _source_selection_reason(
    *,
    is_procurement_domain: bool,
    has_vendor_candidate_for_purchase: bool,
    has_valid_url: bool,
    has_valid_search_price: bool,
    strict_full_match: bool,
) -> str:
    if is_procurement_domain and has_vendor_candidate_for_purchase:
        return "rejected_procurement_domain_vendor_available"
    if is_procurement_domain:
        return "rejected_procurement_domain_no_vendor_candidate"
    if not has_valid_url:
        return "rejected_missing_or_invalid_offer_source_url"
    if not has_valid_search_price:
        return "rejected_missing_search_price"
    if not strict_full_match:
        return "rejected_non_strict_match"
    return "selected_vendor_candidate"


def _read_csv(path: Path) -> list[dict[str, str]]:
    with path.open("r", encoding="utf-8-sig", newline="") as fh:
        return list(csv.DictReader(fh))


def _write_csv(path: Path, rows: list[dict[str, object]]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    if not rows:
        path.write_text("", encoding="utf-8")
        return
    fieldnames = list(rows[0].keys())
    with path.open("w", encoding="utf-8", newline="") as fh:
        writer = csv.DictWriter(fh, fieldnames=fieldnames)
        writer.writeheader()
        writer.writerows(rows)


def _write_xlsx(path: Path, rows: list[dict[str, object]]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    wb = Workbook()
    ws = wb.active
    ws.title = "small_tender_report"

    if rows:
        headers = list(rows[0].keys())
        ws.append(headers)
        url_columns = {h for h in ("market_price_source_url", "offer_source_url") if h in headers}

        for row in rows:
            ws.append([row.get(h) for h in headers])
            current_row = ws.max_row
            for col_name in url_columns:
                value = row.get(col_name)
                if not isinstance(value, str):
                    continue
                value = value.strip()
                if not value or not (value.startswith("http://") or value.startswith("https://")):
                    continue
                col_idx = headers.index(col_name) + 1
                cell = ws.cell(row=current_row, column=col_idx)
                cell.hyperlink = value
                cell.style = "Hyperlink"

    wb.save(path)


def _to_float(value: object) -> float | None:
    if value is None:
        return None
    text = str(value).strip().replace("\xa0", " ").replace(" ", "").replace(",", ".")
    if not text:
        return None
    match = re.search(r"-?\d+(?:\.\d+)?", text)
    if not match:
        return None
    try:
        return float(match.group(0))
    except ValueError:
        return None


def _pick_first_price(row: dict[str, object], fields: list[str]) -> float | None:
    for field in fields:
        price = _to_float(row.get(field))
        if price is not None:
            return price
    return None


def build_single_file_manifest(*, source_csv: Path, attachment_path: Path, out_manifest_csv: Path) -> Path:
    rows = _read_csv(source_csv)
    purchase_ids = sorted({str(r.get("purchase_external_id", "")).strip() for r in rows if str(r.get("purchase_external_id", "")).strip()})
    out_manifest_csv.parent.mkdir(parents=True, exist_ok=True)
    with out_manifest_csv.open("w", encoding="utf-8", newline="") as fh:
        writer = csv.DictWriter(fh, fieldnames=["purchase_external_id", "attachment_path"])
        writer.writeheader()
        for pid in purchase_ids:
            writer.writerow({"purchase_external_id": pid, "attachment_path": str(attachment_path)})
    return out_manifest_csv


def build_manifest_from_attachments_dir(*, source_csv: Path, attachments_root: Path, out_manifest_csv: Path) -> Path:
    rows = _read_csv(source_csv)
    purchase_ids = sorted({str(r.get("purchase_external_id", "")).strip() for r in rows if str(r.get("purchase_external_id", "")).strip()})

    allowed_suffixes = {".pdf", ".doc", ".docx", ".xls", ".xlsx", ".txt", ".rtf", ".zip"}

    def _score(path: Path) -> tuple[int, int, str]:
        name = path.name.lower().replace("ё", "е")
        likely_tz = any(marker in name for marker in ("тз", "тех", "техническ", "spec", "задани"))
        noisy_doc = any(marker in name for marker in ("регистрац", "эп", "инструкц", "пользовател"))
        # Lower score is better.
        return (0 if likely_tz else 1, 1 if noisy_doc else 0, name)

    out_rows: list[dict[str, str]] = []
    for pid in purchase_ids:
        pid_dir = attachments_root / pid
        if not pid_dir.exists() or not pid_dir.is_dir():
            continue
        candidates = [p for p in pid_dir.iterdir() if p.is_file() and p.suffix.lower() in allowed_suffixes]
        if not candidates:
            continue
        for path in sorted(candidates, key=_score):
            out_rows.append({"purchase_external_id": pid, "attachment_path": str(path)})

    out_manifest_csv.parent.mkdir(parents=True, exist_ok=True)
    with out_manifest_csv.open("w", encoding="utf-8", newline="") as fh:
        writer = csv.DictWriter(fh, fieldnames=["purchase_external_id", "attachment_path"])
        writer.writeheader()
        writer.writerows(out_rows)
    return out_manifest_csv
