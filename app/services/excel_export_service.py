from __future__ import annotations

from collections import Counter
from datetime import datetime
from decimal import Decimal
from pathlib import Path
from urllib.parse import urlparse

from openpyxl import Workbook
from openpyxl.formatting.rule import FormulaRule
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from sqlalchemy import select
from sqlalchemy.orm import Session, selectinload

from app.models import (
    BusinessRule,
    CalculationOfferUsage,
    DashboardNotification,
    ExportJob,
    ImportJob,
    ItemAttributes,
    ItemCostCalculation,
    MarketOffer,
    Purchase,
    PurchaseCalculation,
    PurchaseDecisionScore,
    PurchaseWatchlist,
    Supplier,
)
from app.reports.report_builder import build_daily_digest
from app.utils.time import utc_now


def export_to_excel(session: Session, output_path: Path) -> Path:
    output_path.parent.mkdir(parents=True, exist_ok=True)

    wb = Workbook()
    summary_ws = wb.active
    summary_ws.title = "Summary"

    purchases = session.scalars(select(Purchase)).all()
    calculations = session.scalars(select(PurchaseCalculation)).all()
    decision_scores = session.scalars(select(PurchaseDecisionScore)).all()
    watchlist_rows = session.scalars(select(PurchaseWatchlist).order_by(PurchaseWatchlist.updated_at.desc())).all()
    notifications = session.scalars(select(DashboardNotification)).all()
    item_calculations = session.scalars(select(ItemCostCalculation)).all()
    import_jobs = session.scalars(select(ImportJob).order_by(ImportJob.created_at.desc())).all()
    offers = session.scalars(select(MarketOffer)).all()
    suppliers = session.scalars(select(Supplier)).all()
    business_rules = session.scalars(select(BusinessRule).order_by(BusinessRule.key.asc())).all()
    item_attrs = session.scalars(select(ItemAttributes)).all()
    attrs_by_item_id = {row.purchase_item_id: row for row in item_attrs}
    usages = session.scalars(select(CalculationOfferUsage)).all()
    usage_by_offer_id = {usage.market_offer_id: usage for usage in usages if usage.market_offer_id is not None}

    summary_ws.append(["metric", "value"])
    summary_ws.append(["generated_at", utc_now().isoformat()])
    summary_ws.append(["total_purchases", len(purchases)])
    summary_ws.append(["total_purchase_calculations", len(calculations)])
    summary_ws.append(["total_decision_scores", len(decision_scores)])
    summary_ws.append(["total_watchlist_entries", len(watchlist_rows)])
    summary_ws.append(["total_item_calculations", len(item_calculations)])

    source_counter = Counter([purchase.source for purchase in purchases])
    for source, count in sorted(source_counter.items()):
        summary_ws.append([f"purchases_by_source:{source}", count])

    status_counter = Counter([(purchase.status or "unknown") for purchase in purchases])
    for status, count in sorted(status_counter.items()):
        summary_ws.append([f"purchases_by_status:{status}", count])

    filtered_total = sum(job.filtered_count for job in import_jobs)
    summary_ws.append(["total_filtered_purchases", filtered_total])

    latest_import = import_jobs[0].created_at.isoformat() if import_jobs else None
    summary_ws.append(["last_import_at", latest_import])

    no_price_items = sum(1 for calc in item_calculations if calc.status in {"no_relevant_offers", "needs_manual_price_search"})
    insufficient_items = sum(1 for calc in item_calculations if calc.status == "insufficient_market_quantity")
    delivery_unknown_items = sum(1 for calc in item_calculations if "delivery_unknown" in (calc.risk_flags or []))
    used_offers_count = len(usages)

    summary_ws.append(["items_without_prices", no_price_items])
    summary_ws.append(["items_insufficient_market_quantity", insufficient_items])
    summary_ws.append(["items_delivery_unknown", delivery_unknown_items])
    summary_ws.append(["offers_used_in_calculation", used_offers_count])
    summary_ws.append(["notifications_total", len(notifications)])
    summary_ws.append(["notifications_unread", sum(1 for row in notifications if row.status == "unread")])
    summary_ws.append(["strong_recommend_count", sum(1 for row in decision_scores if row.decision == "strong_recommend")])
    summary_ws.append(["recommend_count", sum(1 for row in decision_scores if row.decision == "recommend")])
    summary_ws.append(["needs_manual_review_count", sum(1 for row in decision_scores if row.decision == "needs_manual_review")])
    top_explanations = [calc for calc in calculations if calc.explanation_summary][:10]
    for calc in top_explanations:
        summary_ws.append([f"explanation_purchase_{calc.purchase_id}", calc.explanation_summary])

    purchases_ws = wb.create_sheet("Purchases")
    purchases_ws.append(
        [
            "id",
            "source",
            "external_id",
            "title",
            "url",
            "status",
            "region",
            "region_code",
            "customer_name",
            "submission_deadline",
            "max_total_price",
            "parsed_at",
            "profit_before_tax",
            "profit_after_tax",
            "margin_before_tax_percent",
            "margin_after_tax_percent",
            "tax_amount",
            "vat_amount",
            "data_quality",
            "explanation_summary",
            "risk_level",
            "decision",
            "decision_status",
            "score_total",
            "next_action",
            "decision_reason",
            "watchlist_status",
            "deadline_status",
            "cash_roi_percent",
            "competition_level",
        ]
    )
    calc_by_purchase_id = {c.purchase_id: c for c in calculations}
    score_by_purchase_id = {row.purchase_id: row for row in decision_scores}
    watchlist_by_purchase_id = {row.purchase_id: row for row in watchlist_rows}
    for purchase in purchases:
        calc = calc_by_purchase_id.get(purchase.id)
        score = score_by_purchase_id.get(purchase.id)
        watch = watchlist_by_purchase_id.get(purchase.id)
        purchases_ws.append(
            [
                purchase.id,
                purchase.source,
                purchase.external_id,
                purchase.title,
                purchase.url,
                purchase.status,
                purchase.region,
                purchase.region_code,
                purchase.customer_name,
                purchase.submission_deadline,
                float(purchase.max_total_price) if purchase.max_total_price is not None else None,
                purchase.parsed_at,
                float(calc.profit_before_tax) if calc is not None else None,
                float(calc.profit_after_tax) if calc is not None else None,
                float(calc.margin_before_tax_percent) if calc is not None else None,
                float(calc.margin_after_tax_percent) if calc is not None else None,
                float(calc.tax_amount) if calc is not None else None,
                float(calc.vat_amount) if calc is not None else None,
                purchase.data_quality,
                _decision_explanation_summary(score, calc),
                calc.risk_level if calc is not None else None,
                score.decision if score is not None else None,
                score.decision_status if score is not None else None,
                float(score.score_total) if score is not None else None,
                score.next_action if score is not None else None,
                score.decision_reason if score is not None else None,
                watch.status if watch is not None else None,
                score.deadline_status if score is not None else purchase.deadline_status,
                float(score.cash_roi_percent) if score is not None and score.cash_roi_percent is not None else None,
                score.competition_level if score is not None else None,
            ]
        )

    items_ws = wb.create_sheet("Items")
    items_ws.append(
        [
            "id",
            "purchase_id",
            "purchase_source",
            "purchase_external_id",
            "position_external_id",
            "position_hash",
            "item_name",
            "category",
            "brand",
            "model",
            "article",
            "original_required",
            "compatible_allowed",
            "description",
            "okpd2",
            "quantity",
            "unit",
            "max_unit_price",
            "max_total_price",
            "delivery_region",
            "delivery_address",
            "found_offers_count",
            "relevant_offers_count",
            "used_offers_count",
            "item_cost_status",
            "calculated_total_cost",
            "item_margin_percent",
            "overbuy_quantity",
            "overbuy_cost",
        ]
    )

    purchases_with_items = session.scalars(select(Purchase).options(selectinload(Purchase.items))).all()
    calc_by_item_id = {calc.purchase_item_id: calc for calc in item_calculations}

    offers_by_item: dict[int, list[MarketOffer]] = {}
    for offer in offers:
        if offer.purchase_item_id is not None:
            offers_by_item.setdefault(offer.purchase_item_id, []).append(offer)

    usage_count_by_offer_id = Counter([usage.market_offer_id for usage in usages if usage.market_offer_id is not None])

    for purchase in purchases_with_items:
        for item in purchase.items:
            item_offers = offers_by_item.get(item.id, [])
            found_count = len(item_offers)
            relevant_count = sum(1 for offer in item_offers if offer.is_relevant)
            used_count = sum(usage_count_by_offer_id.get(offer.id, 0) for offer in item_offers)
            calc = calc_by_item_id.get(item.id)
            attrs = attrs_by_item_id.get(item.id)
            calc_total = Decimal(calc.estimated_item_cost) if calc and calc.estimated_item_cost is not None else None
            margin_percent = None
            if calc_total is not None and item.max_total_price is not None and Decimal(item.max_total_price) > 0:
                margin_percent = float(((Decimal(item.max_total_price) - calc_total) / Decimal(item.max_total_price)) * Decimal("100"))

            items_ws.append(
                [
                    item.id,
                    item.purchase_id,
                    purchase.source,
                    purchase.external_id,
                    item.position_external_id,
                    item.position_hash,
                    item.item_name,
                    attrs.category if attrs else None,
                    attrs.brand if attrs else None,
                    attrs.model if attrs else None,
                    attrs.article if attrs else None,
                    bool(attrs.original_required) if attrs else None,
                    bool(attrs.compatible_allowed) if attrs else None,
                    item.description,
                    item.okpd2,
                    float(item.quantity),
                    item.unit,
                    float(item.max_unit_price) if item.max_unit_price is not None else None,
                    float(item.max_total_price) if item.max_total_price is not None else None,
                    item.delivery_region,
                    item.delivery_address,
                    found_count,
                    relevant_count,
                    used_count,
                    calc.status if calc else None,
                    float(calc_total) if calc_total is not None else None,
                    margin_percent,
                    (calc.calculation_details_json or {}).get("overbuy_quantity") if calc else None,
                    (calc.calculation_details_json or {}).get("overbuy_cost") if calc else None,
                ]
            )

    offers_ws = wb.create_sheet("Offers")
    offers_ws.append(
        [
            "id",
            "purchase_external_id",
            "position_external_id",
            "item_name",
            "offer_title",
            "seller_name",
            "region",
            "unit_price",
            "available_quantity",
            "delivery_price",
            "effective_unit_price",
            "margin_percent",
            "relevance_score",
            "is_relevant",
            "supplier_status",
            "match_score",
            "hard_reject_reason",
            "match_reasons",
            "matched_fields",
            "mismatched_fields",
            "delivery_type",
            "delivery_price_type",
            "min_order_quantity",
            "package_quantity",
            "used_in_calculation",
            "taken_quantity",
            "offer_url",
            "risk_flags",
            "tech_spec_confirmation_status",
        ]
    )

    purchase_by_external_id = {purchase.external_id: purchase for purchase in purchases}

    for offer in offers:
        purchase_id = offer.purchase_id
        if purchase_id is None and offer.purchase_external_id:
            purchase = purchase_by_external_id.get(offer.purchase_external_id)
            purchase_id = purchase.id if purchase is not None else None
        calc = calc_by_purchase_id.get(purchase_id) if purchase_id is not None else None
        margin_percent = _resolve_margin_percent(calc)
        tech_spec_confirmation_status = _resolve_tech_spec_confirmation_status(
            is_relevant=offer.is_relevant,
            hard_reject_reason=offer.hard_reject_reason,
            matched_fields=offer.matched_fields_json,
            mismatched_fields=offer.mismatched_fields_json,
            relevance_score=offer.relevance_score,
            match_score=offer.match_score,
            margin_percent=margin_percent,
        )
        if tech_spec_confirmation_status not in {"green", "yellow"}:
            continue
        if not _is_real_source_url(offer.offer_url):
            continue
        if margin_percent is None:
            continue

        usage = usage_by_offer_id.get(offer.id)
        offers_ws.append(
            [
                offer.id,
                offer.purchase_external_id,
                offer.position_external_id,
                offer.item_name,
                offer.offer_title,
                offer.seller_name or offer.supplier_name,
                offer.region,
                float(offer.unit_price),
                offer.available_quantity,
                float(offer.delivery_price) if offer.delivery_price is not None else None,
                float(offer.effective_unit_price) if offer.effective_unit_price is not None else None,
                float(margin_percent),
                float(offer.relevance_score) if offer.relevance_score is not None else None,
                bool(offer.is_relevant),
                offer.supplier_status,
                float(offer.match_score) if offer.match_score is not None else None,
                offer.hard_reject_reason,
                ",".join(offer.match_reasons_json or []),
                ",".join(offer.matched_fields_json or []),
                ",".join(offer.mismatched_fields_json or []),
                offer.delivery_type,
                offer.delivery_price_type,
                offer.min_order_quantity,
                offer.package_quantity,
                bool(usage is not None),
                usage.taken_quantity if usage else None,
                offer.offer_url,
                ",".join(offer.risk_flags or []),
                tech_spec_confirmation_status,
            ]
        )

    calc_ws = wb.create_sheet("Calculations")
    calc_ws.append(
        [
            "purchase_id",
            "max_total_price",
            "estimated_cost",
            "estimated_profit",
            "margin_percent",
            "cash_required",
            "recommendation_status",
            "problematic_items_count",
            "unknown_delivery_items_count",
            "attractiveness_score",
            "cost_before_tax",
            "cost_after_tax",
            "profit_before_tax",
            "profit_after_tax",
            "margin_before_tax_percent",
            "margin_after_tax_percent",
            "vat_amount",
            "tax_amount",
            "explanation_summary",
            "risk_level",
        ]
    )
    for calc in calculations:
        calc_ws.append(
            [
                calc.purchase_id,
                float(calc.max_total_price),
                float(calc.estimated_cost),
                float(calc.estimated_profit),
                float(calc.margin_percent),
                float(calc.cash_required),
                calc.recommendation_status,
                calc.problematic_items_count,
                calc.unknown_delivery_items_count,
                float(calc.attractiveness_score),
                float(calc.cost_before_tax),
                float(calc.cost_after_tax),
                float(calc.profit_before_tax),
                float(calc.profit_after_tax),
                float(calc.margin_before_tax_percent),
                float(calc.margin_after_tax_percent),
                float(calc.vat_amount),
                float(calc.tax_amount),
                calc.explanation_summary,
                calc.risk_level,
            ]
        )

    suppliers_ws = wb.create_sheet("Suppliers")
    suppliers_ws.append(["id", "name", "normalized_name", "status", "rating", "comment"])
    for supplier in suppliers:
        suppliers_ws.append(
            [
                supplier.id,
                supplier.name,
                supplier.normalized_name,
                supplier.status,
                float(supplier.rating) if supplier.rating is not None else None,
                supplier.comment,
            ]
        )

    rules_ws = wb.create_sheet("BusinessRules")
    rules_ws.append(["id", "key", "value", "description", "updated_at"])
    for rule in business_rules:
        rules_ws.append([rule.id, rule.key, rule.value, rule.description, rule.updated_at])

    decisions_ws = wb.create_sheet("DecisionScores")
    decisions_ws.append(
        [
            "purchase_id",
            "decision",
            "risk_level",
            "score_total",
            "score_margin",
            "score_profit",
            "score_deadline",
            "score_data_quality",
            "score_supplier_quality",
            "score_competition",
            "score_cash_efficiency",
            "score_risk",
            "cash_roi_percent",
            "deadline_status",
            "competition_level",
            "decision_reason",
            "next_action",
            "decision_status",
            "explanation_summary",
            "updated_at",
        ]
    )
    for row in decision_scores:
        decisions_ws.append(
            [
                row.purchase_id,
                row.decision,
                row.risk_level,
                float(row.score_total),
                float(row.score_margin),
                float(row.score_profit),
                float(row.score_deadline),
                float(row.score_data_quality),
                float(row.score_supplier_quality),
                float(row.score_competition),
                float(row.score_cash_efficiency),
                float(row.score_risk),
                float(row.cash_roi_percent) if row.cash_roi_percent is not None else None,
                row.deadline_status,
                row.competition_level,
                row.decision_reason,
                row.next_action,
                row.decision_status,
                _decision_explanation_summary(row, calc_by_purchase_id.get(row.purchase_id)),
                row.updated_at,
            ]
        )

    watchlist_ws = wb.create_sheet("Watchlist")
    watchlist_ws.append(["purchase_id", "external_id", "source", "status", "note", "updated_at"])
    purchase_map = {row.id: row for row in purchases}
    for row in watchlist_rows:
        purchase = purchase_map.get(row.purchase_id)
        watchlist_ws.append(
            [
                row.purchase_id,
                purchase.external_id if purchase else None,
                purchase.source if purchase else None,
                row.status,
                row.note,
                row.updated_at,
            ]
        )

    digest = build_daily_digest(session, limit=10)
    digest_ws = wb.create_sheet("DailyDigest")
    digest_ws.append(["metric", "value"])
    digest_ws.append(["generated_at", digest.generated_at])
    digest_ws.append(["new_purchases_24h", digest.new_purchases_24h])
    digest_ws.append(["calculated_24h", digest.calculated_24h])
    digest_ws.append(["strong_recommend_count", digest.strong_recommend_count])
    digest_ws.append(["recommend_count", digest.recommend_count])
    digest_ws.append(["needs_review_count", digest.needs_review_count])
    digest_ws.append([])
    digest_ws.append(["Top opportunities", ""])
    digest_ws.append(
        [
            "external_id",
            "source",
            "max_total_price",
            "profit_after_tax",
            "margin_after_tax_percent",
            "score_total",
            "deadline",
            "next_action",
        ]
    )
    for row in digest.top_rows:
        digest_ws.append(
            [
                row.get("external_id"),
                row.get("source"),
                row.get("max_total_price"),
                row.get("profit_after_tax"),
                row.get("margin_after_tax_percent"),
                row.get("score_total"),
                row.get("deadline"),
                row.get("next_action"),
            ]
        )

    risks_ws = wb.create_sheet("Risks")
    risks_ws.append(
        [
            "purchase_id",
            "external_id",
            "source",
            "decision",
            "risk_level",
            "score_risk",
            "decision_reason",
            "next_action",
            "deadline_status",
        ]
    )
    for row in decision_scores:
        if row.risk_level not in {"high", "critical"}:
            continue
        purchase = purchase_map.get(row.purchase_id)
        risks_ws.append(
            [
                row.purchase_id,
                purchase.external_id if purchase else None,
                purchase.source if purchase else None,
                row.decision,
                row.risk_level,
                float(row.score_risk),
                row.decision_reason,
                row.next_action,
                row.deadline_status,
            ]
        )

    _polish_workbook(wb)
    wb.save(output_path)

    export_job = ExportJob(status="completed", output_path=str(output_path), details={"format": "xlsx"})
    session.add(export_job)
    session.commit()

    return output_path


def _polish_workbook(wb: Workbook) -> None:
    header_fill = PatternFill(fill_type="solid", fgColor="E9EEF5")
    header_font = Font(bold=True)
    link_font = Font(color="0563C1", underline="single")

    for ws in wb.worksheets:
        if ws.max_row < 1 or ws.max_column < 1:
            continue

        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(vertical="center")

        if ws.max_row > 1:
            ws.freeze_panes = "A2"
            ws.auto_filter.ref = ws.dimensions

        for col_idx in range(1, ws.max_column + 1):
            letter = get_column_letter(col_idx)
            values = [ws.cell(row=r, column=col_idx).value for r in range(1, ws.max_row + 1)]
            max_len = max((len(str(v)) for v in values if v is not None), default=8)
            ws.column_dimensions[letter].width = min(max(max_len + 2, 10), 60)

        headers = [str(ws.cell(row=1, column=c).value or "").lower() for c in range(1, ws.max_column + 1)]
        for col_idx, header in enumerate(headers, start=1):
            if "url" not in header:
                continue
            for row in range(2, ws.max_row + 1):
                cell = ws.cell(row=row, column=col_idx)
                if isinstance(cell.value, str) and cell.value.startswith(("http://", "https://")):
                    cell.hyperlink = cell.value
                    cell.font = link_font

        if ws.title == "Purchases":
            _apply_decision_conditional_formatting(ws)


def _decision_explanation_summary(score: PurchaseDecisionScore | None, calc: PurchaseCalculation | None) -> str | None:
    if score and score.explanation_json and score.explanation_json.get("final_reason"):
        return str(score.explanation_json.get("final_reason"))
    if calc:
        return calc.explanation_summary
    return None


def _resolve_margin_percent(calc: PurchaseCalculation | None) -> Decimal | None:
    if calc is None:
        return None
    if calc.margin_percent is not None:
        return Decimal(calc.margin_percent)
    if calc.margin_after_tax_percent is not None:
        return Decimal(calc.margin_after_tax_percent)
    return None


def _is_real_source_url(url: str | None) -> bool:
    if not url:
        return False
    candidate = str(url).strip()
    if not candidate:
        return False

    parsed = urlparse(candidate)
    if parsed.scheme not in {"http", "https"}:
        return False
    hostname = (parsed.hostname or "").lower()
    if not hostname:
        return False
    if hostname == "example.com" or hostname.endswith(".example.com"):
        return False
    return True


def _resolve_tech_spec_confirmation_status(
    *,
    is_relevant: bool | None,
    hard_reject_reason: str | None,
    matched_fields: list[str] | None,
    mismatched_fields: list[str] | None,
    relevance_score: Decimal | float | None,
    match_score: Decimal | float | None,
    margin_percent: Decimal | float | None,
) -> str | None:
    if not is_relevant or hard_reject_reason:
        return "reject"

    key_fields = {"brand", "model", "article", "category", "color", "params", "parameter", "parameters"}
    matched_key = {field for field in (matched_fields or []) if field in key_fields}
    mismatched_key = {field for field in (mismatched_fields or []) if field in key_fields}
    available_key = matched_key | mismatched_key

    if not available_key:
        return "reject"

    margin_value = Decimal(str(margin_percent)) if margin_percent is not None else None
    is_full_match = bool(matched_key) and not mismatched_key and matched_key == available_key
    if is_full_match:
        return "green"

    is_partial_match = bool(matched_key) and bool(mismatched_key)
    if is_partial_match and margin_value is not None and margin_value <= Decimal("30"):
        return "yellow"

    return "reject"


def _apply_decision_conditional_formatting(ws) -> None:
    decision_col = None
    status_col = None
    for col_idx in range(1, ws.max_column + 1):
        title = str(ws.cell(row=1, column=col_idx).value or "").strip().lower()
        if title == "decision":
            decision_col = col_idx
        if title == "decision_status":
            status_col = col_idx

    if decision_col is None or ws.max_row < 2:
        return

    letter = get_column_letter(decision_col)
    data_range = f"{letter}2:{letter}{ws.max_row}"
    strong_fill = PatternFill(fill_type="solid", fgColor="C6EFCE")
    review_fill = PatternFill(fill_type="solid", fgColor="FFEB9C")
    reject_fill = PatternFill(fill_type="solid", fgColor="FFC7CE")

    ws.conditional_formatting.add(
        data_range,
        FormulaRule(formula=[f'OR(LOWER({letter}2)="strong_recommend",LOWER({letter}2)="strong_buy")'], fill=strong_fill),
    )
    ws.conditional_formatting.add(
        data_range,
        FormulaRule(formula=[f'LOWER({letter}2)="needs_manual_review"'], fill=review_fill),
    )
    ws.conditional_formatting.add(
        data_range,
        FormulaRule(formula=[f'LOWER({letter}2)="reject"'], fill=reject_fill),
    )

    if status_col is None:
        return
    status_letter = get_column_letter(status_col)
    status_range = f"{status_letter}2:{status_letter}{ws.max_row}"
    ws.conditional_formatting.add(
        status_range,
        FormulaRule(formula=[f'LOWER({status_letter}2)="needs_review"'], fill=review_fill),
    )
